from collections import defaultdict
from datetime import date, time, timedelta
import datetime
from doctest import debug
import json
from zoneinfo import ZoneInfo                   
from django.urls import reverse
from django.utils import timezone
import os
import random
import uuid
import requests
import openpyxl
from datetime import datetime as dt
from core.models import SiteSetting
from core.utils.mailers import send_account_email_simple, send_student_rfid_account_email
from django.views.decorators.http import require_POST
from datetime import date as ddate, datetime as dt
import core.utils.mailers
from django.db import connection, transaction
from django.utils.http import url_has_allowed_host_and_scheme
from django.views.decorators.http import require_http_methods
from django.utils.html import strip_spaces_between_tags as minify_html
from django.template.loader import render_to_string
from PyPDF2 import PdfMerger
import io, zipfile, datetime, base64

from teacher import realtime
from .forms import SiteSettingForm
from .utils.roles import is_admin_role


from django.contrib import messages
from django.contrib.auth.hashers import make_password, check_password
from django.db import IntegrityError, connection
from django.http import HttpResponse, HttpResponseBadRequest, HttpResponseNotAllowed, JsonResponse
from django.shortcuts import render, redirect
from django.contrib.auth import logout
from django.views.decorators.csrf import csrf_exempt
from django.core.mail import send_mail
from django.conf import settings
from django.db import connection, ProgrammingError
from django.views.decorators.http import require_POST, require_GET


from openpyxl.utils import get_column_letter
from django.conf import settings
from django.db import connection


from teacher.views import _fmt_hhmm
try:
    from weasyprint import HTML, CSS
    WEASY_OK = True
except Exception:
    WEASY_OK = False

    from requests_ntlm import HttpNtlmAuth

def _make_auth_headers_and_authobj():
    user = settings.SSRS_NTLM_USER
    password = settings.SSRS_NTLM_PASS
    auth = HttpNtlmAuth(user, password)
    headers = {}  # Add custom headers if needed
    return headers, auth


def build_user_header_context(user_id: int):
    """
    Unified header context for any FACULTY (admin/teacher).
    Provides both admin_* and teacher_* keys for legacy templates.
    Counts unread notifications where status is 'Unread' or 'Unseen' (case-insensitive).
    """
    # default skeleton (no user/session)
    ctx = {
        "admin_name": "Admin",
        "admin_profile_image": None,
        "teacher_name": "Teacher",
        "teacher_profile_image": None,
        "notifications": [],
        "unread_count": 0,
        "MEDIA_URL": settings.MEDIA_URL,

        "PUSHER_KEY": getattr(settings, "PUSHER_KEY", ""),
        "PUSHER_CLUSTER": getattr(settings, "PUSHER_CLUSTER", ""),
    }

    if not user_id:
        return ctx

    # name + avatar
    with connection.cursor() as c:
        c.execute("""
            SELECT first_name, middle_name, last_name, extension_name, profile_image, ISNULL(is_archived,0)
            FROM FACULTY
            WHERE faculty_id = %s
        """, [user_id])
        row = c.fetchone()

    if row:
        first, middle, last, ext, profile_image, is_archived = row
        middle_part = f" {middle[0]}." if middle else ""
        ext_part = f" {ext}" if ext else ""
        name = f"{(first or '').strip()}{middle_part} {(last or '').strip()}{ext_part}".strip().title()
        if is_archived:
            name += " (Inactive)"
        ctx["admin_name"] = name
        ctx["teacher_name"] = name
        ctx["admin_profile_image"] = profile_image
        ctx["teacher_profile_image"] = profile_image

    # notifications for this user
    with connection.cursor() as c:
        c.execute("""
            SELECT notification_id, message, status, created_at
            FROM NOTIFICATIONS
            WHERE receiver_teacher_id = %s
            ORDER BY created_at DESC
        """, [user_id])
        notif_rows = c.fetchall() or []

    notifications = [
        {"id": r[0], "message": r[1], "status": r[2], "created_at": r[3]}
        for r in notif_rows
    ]
    ctx["notifications"] = notifications

    # unread → 'Unread' or 'Unseen' (case-insensitive)
    ctx["unread_count"] = sum(
        1 for n in notifications
        if str(n.get("status", "")).strip().lower() in ("unread", "unseen")
    )

    return ctx

def is_rfid_duplicate(rfid, exclude_student_id=None, exclude_faculty_id=None):
    """
    Returns True if this RFID is already assigned to ANY active student or faculty.
    You can optionally exclude a specific student/faculty ID (for edit cases).
    """
    if not rfid:
        return False

    # --- Check STUDENTS ---
    with connection.cursor() as c:
        sql = """
            SELECT COUNT(*) 
            FROM STUDENTS 
            WHERE rfid_sticker_id = %s 
              AND ISNULL(is_archived,0) = 0
        """
        params = [rfid]
        if exclude_student_id:
            sql += " AND student_id <> %s"
            params.append(exclude_student_id)

        c.execute(sql, params)
        if (c.fetchone() or [0])[0] > 0:
            return True

    # --- Check FACULTY ---
    with connection.cursor() as c:
        sql = """
            SELECT COUNT(*) 
            FROM FACULTY 
            WHERE rfid_sticker_id = %s 
              AND ISNULL(is_archived,0) = 0
        """
        params = [rfid]
        if exclude_faculty_id:
            sql += " AND faculty_id <> %s"
            params.append(exclude_faculty_id)

        c.execute(sql, params)
        if (c.fetchone() or [0])[0] > 0:
            return True

    return False


def get_admin_header_context(admin_id):
    with connection.cursor() as cursor:
        # Name + avatar
        cursor.execute("""
            SELECT first_name, middle_name, last_name, extension_name, profile_image
            FROM FACULTY
            WHERE faculty_id = %s
        """, [admin_id])
        row = cursor.fetchone()

    if row:
        first, middle, last, ext, profile_image = row
        middle_part = f" {middle[0]}." if middle else ""
        ext_part = f" {ext}" if ext else ""
        name = f"{first}{middle_part} {last}{ext_part}".strip().title()
    else:
        name, profile_image = "Admin", None

    # Notifications (keep your schema)
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT notification_id, message, status, created_at
            FROM NOTIFICATIONS
            WHERE receiver_teacher_id = %s
            ORDER BY created_at DESC
        """, [admin_id])
        notifs = cursor.fetchall()

    notifications = [
        {"id": n[0], "message": n[1], "status": n[2], "created_at": n[3]}
        for n in notifs
    ]
    unread_count = sum(1 for n in notifications if str(n["status"]).lower() == "unread")

    # Return BOTH admin_* and teacher_* so old templates keep working
    return {
        "admin_name": name,
        "admin_profile_image": profile_image,
        "teacher_name": name,                 # alias for legacy templates
        "teacher_profile_image": profile_image,  # alias for legacy templates
        "notifications": notifications,
        "unread_count": unread_count,
        "MEDIA_URL": settings.MEDIA_URL,
    }
def _none_if_empty(v):
    if v is None:
        return None
    s = str(v).strip()
    return None if s == "" or s.lower() == "none" else v

def _int_or_none(v):
    v = _none_if_empty(v)
    if v is None:
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        return None

def _date_or_none(v):
    return _none_if_empty(v)

# ---------- auth helper ----------
def _make_auth_headers_and_authobj():
    mode = getattr(settings, "SSRS_AUTH_MODE", "NONE").upper()
    if mode == "NTLM":
        try:
            from requests_ntlm import HttpNtlmAuth
        except ImportError:
            raise RuntimeError("requests-ntlm is not installed. Run: pip install requests-ntlm")
        user = getattr(settings, "SSRS_NTLM_USER", "")
        pwd  = getattr(settings, "SSRS_NTLM_PASS", "")
        return ({}, HttpNtlmAuth(user, pwd))
    elif mode == "BASIC":
        user = getattr(settings, "SSRS_BASIC_USER", "")
        pwd  = getattr(settings, "SSRS_BASIC_PASS", "")
        token = base64.b64encode(f"{user}:{pwd}".encode("utf-8")).decode("ascii")
        return ({"Authorization": f"Basic {token}"}, None)
    else:
        return ({}, None)
    
def _normalize_time_label(v):
    if v is None:
        return ""
    s = str(v).strip()
    low = s.lower()
    if low == "midnight":
        return "12:00 a.m."
    if low == "noon":
        return "12:00 p.m."
    return s

# ---------- SSRS fetch (Render) ----------
def _ssrs_pdf(report_path, params, timeout=60):
    base = settings.SSRS_BASE_URL.rstrip('/')
    qp = [("rs:Command", "Render"), ("rs:Format", "PDF")]
    for k, v in params.items():
        if v is not None:
            qp.append((k, str(v)))

    query = "&".join(
        f"{requests.utils.quote(k, safe=':')}={requests.utils.quote(v)}"
        for k, v in qp
    )
    url = f"{base}?{requests.utils.quote(report_path, safe='/')}&{query}"

    headers, auth = _make_auth_headers_and_authobj()

    last = None
    for _ in (1, 2):  # 1 retry on 5xx
        try:
            r = requests.get(
                url, auth=auth, headers=headers,
                verify=getattr(settings, "SSRS_VERIFY_TLS", False),
                timeout=timeout
            )
        except requests.RequestException as ex:
            raise requests.HTTPError(f"SSRS request failed for {url}\n{ex}") from ex

        if r.status_code == 200:
            return r.content
        last = r

        # dump first part of SSRS error to your console for quick diagnosis
        print("SSRS status:", r.status_code, r.reason)
        print("SSRS body:", r.text[:1500])

        if r.status_code < 500:
            break  # don’t retry 4xx etc.

    raise requests.HTTPError(f"{last.status_code} {last.reason} for {url}\n\n{last.text[:4000]}")


# ---------- DB helper ----------
def _filtered_utilization_ids(lab_id, teacher_id, start_date, end_date):
    with connection.cursor() as cursor:
        q = ["SELECT u.utilization_id FROM UTILIZATION_SLIP u WHERE u.status = 'Completed'"]
        p = []
        if lab_id is not None:
            q.append("AND u.lab_id = %s")
            p.append(lab_id)
        if teacher_id is not None:
            q.append("AND u.requested_by = %s")
            p.append(teacher_id)
        if start_date is not None:
            # compare only the date part so the whole start_date is included
            q.append("AND CAST(u.date AS date) >= %s")
            p.append(start_date)
        if end_date is not None:
            # compare only the date part so the whole end_date is included
            q.append("AND CAST(u.date AS date) <= %s")
            p.append(end_date)

        q.append("ORDER BY u.date DESC, u.utilization_id DESC")
        cursor.execute(" ".join(q), p)
        return [row[0] for row in cursor.fetchall()]

# ---------- Dashboard ----------
def dashboard(request):
    if not request.session.get('user_id') or request.session.get('role') != 'admin':
        return redirect('login')

    try:
        context = build_user_header_context(request.session.get('user_id'))
    except Exception:
        context = {}

    lab_id     = _int_or_none(request.GET.get('lab_id'))
    teacher_id = _int_or_none(request.GET.get('teacher_id'))
    start_date = _date_or_none(request.GET.get('start_date'))
    end_date   = _date_or_none(request.GET.get('end_date'))

    with connection.cursor() as cursor:
        cursor.execute("SELECT lab_id, lab_num FROM LABORATORIES")
        labs = cursor.fetchall()

        cursor.execute("""
            SELECT faculty_id, CONCAT(first_name, ' ', last_name)
            FROM FACULTY
            WHERE is_archived = 0
            ORDER BY last_name, first_name
        """)
        teachers = cursor.fetchall()

        query = """
            SELECT
                u.date,
                l.lab_num,
                CONCAT(f.first_name, ' ', f.last_name) AS teacher_name,
                c.course_name,
                u.student_year_and_section,
                u.time_duration,
                (
                    SELECT COUNT(*)
                    FROM COMPUTER_LAB_ATTENDANCE a
                    WHERE a.utilization_id = u.utilization_id
                ) AS students_present,
                u.utilization_id,
                s.term,
                s.school_year
            FROM UTILIZATION_SLIP u
            JOIN LABORATORIES    l  ON u.lab_id     = l.lab_id
            JOIN FACULTY         f  ON u.requested_by = f.faculty_id
            JOIN LAB_SCHEDULE    ls ON u.schedule_id = ls.schedule_id
            JOIN ASSIGNED_TEACHER at ON ls.assigned_teacher_id = at.assigned_teacher_id
            JOIN COURSE          c  ON at.course_id = c.course_id
            JOIN SEMESTER        s  ON at.semester_id = s.semester_id
            WHERE u.status = 'Completed'
        """
        params = []
        if lab_id is not None:
            query += " AND u.lab_id = %s"
            params.append(lab_id)
        if teacher_id is not None:
            query += " AND u.requested_by = %s"
            params.append(teacher_id)
        if start_date is not None:
            query += " AND CAST(u.date AS date) >= %s"
            params.append(start_date)
        if end_date is not None:
            query += " AND CAST(u.date AS date) <= %s"
            params.append(end_date)

        query += " ORDER BY u.date DESC"
        cursor.execute(query, params)
        records = cursor.fetchall()

    selected_lab_val = str(lab_id) if lab_id is not None else ''
    selected_teacher_val = str(teacher_id) if teacher_id is not None else ''
    start_date_val = start_date if start_date is not None else ''
    end_date_val = end_date if end_date is not None else ''

    context.update({
        'current_page': 'Attendance Record',
        'labs': labs,
        'teachers': teachers,
        'selected_lab_id': selected_lab_val,
        'selected_teacher_id': selected_teacher_val,
        'start_date': start_date_val,
        'end_date': end_date_val,
        'attendance_records': [
            {
                'date': r[0],
                'lab_name': r[1],
                'teacher_name': r[2],
                'course': r[3],
                'section': r[4],
                'time_duration': _normalize_time_label(r[5]),
                'students_present': r[6],
                'utilization_id': r[7],
                'semester_text': f"{r[8]}, A.Y. {r[9]}" if r[8] and r[9] else "—",
            }
            for r in records
        ],
    })
    return render(request, 'admin/dashboard.html', context)




# ---------- Single PDFs (row buttons) ----------
def print_utilization_slip(request, utilization_id):
    try:
        pdf = _ssrs_pdf(
            settings.SSRS_UTILIZATION_REPORT_PATH,
            {settings.SSRS_UTIL_PARAM_NAME: utilization_id}
        )
    except requests.HTTPError as e:
        return HttpResponse(f"SSRS error: {e}", status=502)
    resp = HttpResponse(pdf, content_type="application/pdf")
    resp['Content-Disposition'] = f'inline; filename=UtilizationSlip_{utilization_id}.pdf'
    return resp

def print_attendance_sheet(request, utilization_id):
    try:
        pdf = _ssrs_pdf(
            settings.SSRS_ATTENDANCE_REPORT_PATH,
            {settings.SSRS_ATTEND_PARAM_NAME: utilization_id}
        )
    except requests.HTTPError as e:
        return HttpResponse(f"SSRS error: {e}", status=502)
    resp = HttpResponse(pdf, content_type="application/pdf")
    resp['Content-Disposition'] = f'inline; filename=AttendanceSheet_{utilization_id}.pdf'
    return resp

# ---------- Preview modal API ----------
def attendance_preview_api(request):
    uid = request.GET.get("utilization_id")
    if not uid:
        return JsonResponse({"error":"missing utilization_id"}, status=400)

    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT u.date, u.time_duration, l.lab_num,
                   CONCAT(f.first_name,' ',f.last_name) AS instructor,
                   c.course_code, u.student_year_and_section,
                   COALESCE(u.remarks,'') AS remarks,
                   u.processed_by,
                   s.term, s.school_year
            FROM UTILIZATION_SLIP u
            JOIN LABORATORIES l ON u.lab_id = l.lab_id
            JOIN LAB_SCHEDULE ls ON u.schedule_id = ls.schedule_id
            JOIN ASSIGNED_TEACHER at ON ls.assigned_teacher_id = at.assigned_teacher_id
            JOIN COURSE c ON at.course_id = c.course_id
            JOIN FACULTY f ON u.requested_by = f.faculty_id AND f.is_archived = 0
            JOIN SEMESTER s ON at.semester_id = s.semester_id
            WHERE u.utilization_id = %s
        """, [uid])
        row = cursor.fetchone()
        if not row:
            return JsonResponse({"error":"not_found"}, status=404)

        cursor.execute("""
            SELECT CONCAT(s.first_name, ' ',
                          COALESCE(LEFT(s.middle_name,1)+'. ',''),
                          s.last_name) AS full_name
            FROM COMPUTER_LAB_ATTENDANCE a
            JOIN STUDENTS s ON a.student_id = s.student_id
            WHERE a.utilization_id = %s AND s.is_archived = 0
            ORDER BY full_name
        """, [uid])
        students = [r[0] for r in cursor.fetchall()]

        processed_by = ""
        if row[7]:
            cursor.execute("""
                SELECT CONCAT(first_name,' ',last_name)
                FROM FACULTY WHERE faculty_id=%s AND is_archived=0
            """, [row[7]])
            p = cursor.fetchone()
            processed_by = p[0] if p else ""

    return JsonResponse({
        "date": str(row[0]),
        "time": _normalize_time_label(row[1]),
        "lab": row[2],
        "instructor": row[3],
        "course": row[4],
        "section": row[5],
        "remarks": row[6],
        "processed_by": processed_by,
        "semester": f"{row[8]}, A.Y. {row[9]}",
        "students_present": len(students),
        "students": students,
    })

# ---------- Bulk: merged / zip ----------
def print_queue(request):
    t         = request.GET.get("type", "all")  # all | utilization | attendance
    lab_id    = _int_or_none(request.GET.get('lab_id'))
    teacher_id= _int_or_none(request.GET.get('teacher_id'))
    start_date= _date_or_none(request.GET.get('start_date'))
    end_date  = _date_or_none(request.GET.get('end_date'))

    ids = _filtered_utilization_ids(lab_id, teacher_id, start_date, end_date)
    if not ids:
        return HttpResponse("No records to print for the selected filters.", status=404)

    merger = PdfMerger()
    try:
        for uid in ids:
            if t in ("all", "utilization"):
                merger.append(io.BytesIO(_ssrs_pdf(
                    settings.SSRS_UTILIZATION_REPORT_PATH,
                    {settings.SSRS_UTIL_PARAM_NAME: uid}
                )))
            if t in ("all", "attendance"):
                merger.append(io.BytesIO(_ssrs_pdf(
                    settings.SSRS_ATTENDANCE_REPORT_PATH,
                    {settings.SSRS_ATTEND_PARAM_NAME: uid}
                )))
    except requests.HTTPError as e:
        return HttpResponse(f"SSRS error: {e}", status=502)

    out = io.BytesIO(); merger.write(out); merger.close()
    resp = HttpResponse(out.getvalue(), content_type="application/pdf")
    resp['Content-Disposition'] = 'inline; filename=print_queue.pdf'
    return resp

def export_merged_pdf(request):
    t         = request.GET.get("type", "all")
    lab_id    = _int_or_none(request.GET.get('lab_id'))
    teacher_id= _int_or_none(request.GET.get('teacher_id'))
    start_date= _date_or_none(request.GET.get('start_date'))
    end_date  = _date_or_none(request.GET.get('end_date'))

    ids = _filtered_utilization_ids(lab_id, teacher_id, start_date, end_date)
    if not ids:
        return HttpResponse("No records to export for the selected filters.", status=404)

    merger = PdfMerger()
    try:
        for uid in ids:
            if t in ("all", "utilization"):
                merger.append(io.BytesIO(_ssrs_pdf(
                    settings.SSRS_UTILIZATION_REPORT_PATH,
                    {settings.SSRS_UTIL_PARAM_NAME: uid}
                )))
            if t in ("all", "attendance"):
                merger.append(io.BytesIO(_ssrs_pdf(
                    settings.SSRS_ATTENDANCE_REPORT_PATH,
                    {settings.SSRS_ATTEND_PARAM_NAME: uid}
                )))
    except requests.HTTPError as e:
        return HttpResponse(f"SSRS error: {e}", status=502)

    out = io.BytesIO(); merger.write(out); merger.close()
    filename = f"export_{t}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    resp = HttpResponse(out.getvalue(), content_type="application/pdf")
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp

def export_pdfs(request):
    t         = request.GET.get("type", "all")
    lab_id    = _int_or_none(request.GET.get('lab_id'))
    teacher_id= _int_or_none(request.GET.get('teacher_id'))
    start_date= _date_or_none(request.GET.get('start_date'))
    end_date  = _date_or_none(request.GET.get('end_date'))

    ids = _filtered_utilization_ids(lab_id, teacher_id, start_date, end_date)
    if not ids:
        return HttpResponse("No records to export for the selected filters.", status=404)

    buf = io.BytesIO()
    try:
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for uid in ids:
                if t in ("all", "utilization"):
                    zf.writestr(f"UtilizationSlip_{uid}.pdf",
                                _ssrs_pdf(settings.SSRS_UTILIZATION_REPORT_PATH,
                                          {settings.SSRS_UTIL_PARAM_NAME: uid}))
                if t in ("all", "attendance"):
                    zf.writestr(f"AttendanceSheet_{uid}.pdf",
                                _ssrs_pdf(settings.SSRS_ATTENDANCE_REPORT_PATH,
                                          {settings.SSRS_ATTEND_PARAM_NAME: uid}))
    except requests.HTTPError as e:
        return HttpResponse(f"SSRS error: {e}", status=502)

    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    resp = HttpResponse(buf.getvalue(), content_type="application/zip")
    resp['Content-Disposition'] = f'attachment; filename=exports_{t}_{ts}.zip'
    return resp

def print_queue_single(request):
    uid = request.GET.get("utilization_id")
    t   = request.GET.get("type", "both")   # both | utilization | attendance
    if not uid:
        return HttpResponse("Missing utilization_id", status=400)

    merger = PdfMerger()
    try:
        if t in ("both", "utilization"):
            merger.append(io.BytesIO(_ssrs_pdf(
                settings.SSRS_UTILIZATION_REPORT_PATH,
                {settings.SSRS_UTIL_PARAM_NAME: uid}
            )))
        if t in ("both", "attendance"):
            merger.append(io.BytesIO(_ssrs_pdf(
                settings.SSRS_ATTENDANCE_REPORT_PATH,
                {settings.SSRS_ATTEND_PARAM_NAME: uid}
            )))
    except requests.HTTPError as e:
        return HttpResponse(f"SSRS error: {e}", status=502)

    out = io.BytesIO(); merger.write(out); merger.close()
    resp = HttpResponse(out.getvalue(), content_type="application/pdf")
    resp['Content-Disposition'] = f'inline; filename=session_{uid}.pdf'
    return resp
# MANAGE COURSES


def _has_table(table):
    with connection.cursor() as c:
        try:
            c.execute("""
                SELECT 1
                FROM INFORMATION_SCHEMA.TABLES
                WHERE TABLE_NAME = %s
            """, [table])
            return c.fetchone() is not None
        except Exception:
            return False


def _has_column(table, column):
    with connection.cursor() as c:
        try:
            c.execute("""
                SELECT 1
                FROM INFORMATION_SCHEMA.COLUMNS
                WHERE TABLE_NAME = %s AND COLUMN_NAME = %s
            """, [table, column])
            return c.fetchone() is not None
        except Exception:
            try:
                c.execute(f"SELECT TOP 0 [{column}] FROM {table}")
                return True
            except Exception:
                return False


def _parse_date(s):
    if not s:
        return None
    return dt.strptime(s, "%Y-%m-%d").date()


def _safe_redirect_back(request, fallback_name="manage_courses"):
    ref = request.META.get("HTTP_REFERER")
    if ref:
        return redirect(ref)
    try:
        return redirect(fallback_name)
    except Exception:
        return redirect("manage_courses")


def get_admin_header_context(user_id):
    return {
        "teacher_name": "Admin",
        "teacher_profile_image": None,
        "notifications": [],
        "unread_count": 0,
        "current_page": "Manage Courses",
    }

# ---------------- page ----------------
def manage_courses(request):
    """
    Active view (?archived not present): show ONLY active rows.
    Archived view (?archived=1): show ONLY inactive rows.
    """
    if not request.session.get('user_id') or request.session.get('role') != 'admin':
        return redirect('login')

    admin_id = request.session.get('user_id')
    context = build_user_header_context(admin_id)

    q = (request.GET.get('q') or '').strip()
    archived_only = (request.GET.get('archived', '').lower() in ('1', 'true', 'yes'))

    has_is_active_course = _has_column('COURSE', 'is_active')
    has_is_active_sem = _has_column('SEMESTER', 'is_active')
    has_course_program = _has_column('COURSE', 'program_id')
    has_program_table = _has_table('PROGRAM')

    # ----- COURSES -----
    where = []
    params = []
    if q:
        like = f"%{q}%"
        where.append("(c.course_code LIKE %s OR c.course_name LIKE %s)")
        params.extend([like, like])

    if has_is_active_course:
        where.append("c.is_active = %s")
        params.append(0 if archived_only else 1)

    where_sql = f"WHERE {' AND '.join(where)}" if where else ""

    with connection.cursor() as c:
        if has_course_program and has_program_table:
            c.execute(f"""
                SELECT c.course_id, c.course_code, c.course_name, c.program_id,
                       p.program_code, p.program_name
                       {", c.is_active" if has_is_active_course else ""}
                FROM COURSE c
                LEFT JOIN PROGRAM p ON p.program_id = c.program_id
                {where_sql}
                ORDER BY c.course_code
            """, params)
        else:
            c.execute(f"""
                SELECT c.course_id, c.course_code, c.course_name
                       {", c.is_active" if has_is_active_course else ""}
                FROM COURSE c
                {where_sql}
                ORDER BY c.course_code
            """, params)
        course_rows = c.fetchall()

        has_at_is_active = _has_column('ASSIGNED_TEACHER', 'is_active')
        if has_at_is_active:
            c.execute("""
                SELECT course_id, COUNT(*)
                FROM ASSIGNED_TEACHER
                WHERE is_active=1
                GROUP BY course_id
            """)
        else:
            c.execute("SELECT course_id, COUNT(*) FROM ASSIGNED_TEACHER GROUP BY course_id")
        assign_counts = dict(c.fetchall() or [])

        c.execute("""
            SELECT at.course_id, COUNT(*)
            FROM LAB_SCHEDULE s
            JOIN ASSIGNED_TEACHER at ON at.assigned_teacher_id = s.assigned_teacher_id
            WHERE s.date >= CAST(GETDATE() AS DATE)
            GROUP BY at.course_id
        """)
        future_sched_counts = dict(c.fetchall() or [])

        if _has_column('UTILIZATION_SLIP', 'course_id'):
            c.execute("SELECT course_id, COUNT(*) FROM UTILIZATION_SLIP GROUP BY course_id")
            slip_counts = dict(c.fetchall() or [])
        else:
            slip_counts = {}

    courses = []
    for row in course_rows:
        if has_course_program and has_program_table and has_is_active_course:
            cid, code, name, pid, pcode, pname, is_active = row
        elif has_course_program and has_program_table:
            cid, code, name, pid, pcode, pname = row
            is_active = 1
        elif has_is_active_course:
            cid, code, name, is_active = row
            pid = pcode = pname = None
        else:
            cid, code, name = row
            is_active = 1
            pid = pcode = pname = None

        assigned_now = int(assign_counts.get(cid, 0) or 0)
        future_sched = int(future_sched_counts.get(cid, 0) or 0)
        slip_refs = int(slip_counts.get(cid, 0) or 0)

        unused = (assigned_now == 0 and future_sched == 0 and slip_refs == 0)
        can_hard_delete = (not has_is_active_course or is_active == 0) or unused

        courses.append({
            "id": cid,
            "code": code,
            "name": name,
            "program_id": pid,
            "program_code": pcode,
            "program_name": pname,
            "is_active": bool(is_active),
            "can_hard_delete": can_hard_delete,
        })

    # ----- SEMESTERS -----
    with connection.cursor() as c:
        sem_where = []
        sem_params = []
        if has_is_active_sem:
            sem_where.append("is_active = %s")
            sem_params.append(0 if archived_only else 1)
        sem_sql = f"WHERE {' AND '.join(sem_where)}" if sem_where else ""

        c.execute(f"""
            SELECT semester_id, term, school_year, start_date, end_date
                   {", is_active" if has_is_active_sem else ""}
            FROM SEMESTER
            {sem_sql}
            ORDER BY start_date DESC, semester_id DESC
        """, sem_params)
        sem_rows = c.fetchall()

        c.execute("SELECT semester_id, COUNT(*) FROM ASSIGNED_TEACHER GROUP BY semester_id")
        sem_assign_counts = dict(c.fetchall() or [])

        try:
            c.execute("SELECT semester_id, COUNT(*) FROM UTILIZATION_SLIP GROUP BY semester_id")
            sem_util_counts = dict(c.fetchall() or [])
        except Exception:
            sem_util_counts = {}

    today = ddate.today()
    semesters = []
    for row in sem_rows:
        if has_is_active_sem:
            sid, term, sy, sd, ed, is_active = row
        else:
            sid, term, sy, sd, ed = row
            is_active = 1

        if sd and ed and sd <= today <= ed:
            status = "Active"
        elif sd and sd > today:
            status = "Upcoming"
        else:
            status = "Past"

        in_assign = int(sem_assign_counts.get(sid, 0) or 0)
        in_util = int(sem_util_counts.get(sid, 0) or 0)
        unused_sem = (in_assign == 0 and in_util == 0)
        can_hard_delete = (not has_is_active_sem or is_active == 0) or unused_sem

        semesters.append({
            "id": sid,
            "term": term or "",
            "school_year": sy or "",
            "start_date": sd,
            "end_date": ed,
            "status": status,
            "is_active": bool(is_active),
            "can_hard_delete": can_hard_delete,
        })

    # ----- PROGRAMS -----
    programs = []
    programs_all_for_select = []
    has_prog_table = _has_table('PROGRAM')
    has_prog_is_active = _has_column('PROGRAM', 'is_active')

    if has_prog_table:
        with connection.cursor() as c:
            c.execute(f"""
                SELECT program_id, program_code, program_name
                       {", is_active" if has_prog_is_active else ""}
                FROM PROGRAM
                ORDER BY program_code
            """)
            prog_rows_all = c.fetchall()

            if _has_column('COURSE', 'program_id'):
                c.execute("""
                    SELECT program_id, COUNT(*)
                    FROM COURSE
                    WHERE program_id IS NOT NULL
                    GROUP BY program_id
                """)
                prog_course_counts = dict(c.fetchall() or [])
            else:
                prog_course_counts = {}

        for row in prog_rows_all:
            if has_prog_is_active:
                pid, pcode, pname, pactive = row
                pactive = bool(pactive)
            else:
                pid, pcode, pname = row
                pactive = True

            programs_all_for_select.append({
                "id": pid, "code": pcode, "name": pname, "active": pactive
            })

            if has_prog_is_active:
                if archived_only and pactive:
                    continue
                if not archived_only and not pactive:
                    continue

            ref_cnt = int(prog_course_counts.get(pid, 0) or 0)
            unused_prog = (ref_cnt == 0)
            can_hard_delete = (not has_prog_is_active or not pactive) or unused_prog

            programs.append({
                "id": pid,
                "code": pcode,
                "name": pname,
                "active": pactive,
                "ref_count": ref_cnt,
                "can_hard_delete": can_hard_delete,
            })

    context.update({
        "q": q,
        "archived_only": archived_only,
        "courses": courses,
        "semesters": semesters,
        "programs": programs,
        "programs_all_for_select": programs_all_for_select,
        "has_is_active_course": has_is_active_course,
        "has_is_active_sem": has_is_active_sem,
        "current_page": "Manage Courses",
    })
    return render(request, "admin/manage_courses.html", context)


# ---------------- COURSES ----------------

@transaction.atomic
def add_course(request):
    if request.method != "POST":
        return _safe_redirect_back(request, "manage_courses")

    code = (request.POST.get("course_code") or "").strip()
    name = (request.POST.get("course_name") or "").strip()
    program_id = request.POST.get("program_id") or None
    if program_id == "":
        program_id = None

    if not code or not name:
        messages.error(request, "Course code and name are required.")
        return _safe_redirect_back(request, "manage_courses")

    has_is_active = _has_column('COURSE', 'is_active')
    has_program = _has_column('COURSE', 'program_id')

    with connection.cursor() as c:
        try:
            c.execute("SELECT COUNT(*) FROM COURSE WHERE course_code=%s", [code])
            if (c.fetchone() or (0,))[0] > 0:
                messages.error(request, "Course code already exists.")
                return _safe_redirect_back(request, "manage_courses")
        except Exception:
            pass

        if has_is_active and has_program:
            c.execute("""
                INSERT INTO COURSE (course_code, course_name, program_id, is_active)
                VALUES (%s, %s, %s, 1)
            """, [code, name, program_id])
        elif has_is_active:
            c.execute("INSERT INTO COURSE (course_code, course_name, is_active) VALUES (%s, %s, 1)", [code, name])
        elif has_program:
            c.execute("INSERT INTO COURSE (course_code, course_name, program_id) VALUES (%s, %s, %s)", [code, name, program_id])
        else:
            c.execute("INSERT INTO COURSE (course_code, course_name) VALUES (%s, %s)", [code, name])

    messages.success(request, "Course created.")
    return _safe_redirect_back(request, "manage_courses")


@transaction.atomic
def edit_course(request, course_id):
    if request.method != "POST":
        return HttpResponse("Invalid access", status=400)

    code = (request.POST.get("course_code") or "").strip()
    name = (request.POST.get("course_name") or "").strip()
    program_id = request.POST.get("program_id") or None
    if program_id == "":
        program_id = None
    status = (request.POST.get("status") or "active").strip().lower()

    if not code or not name:
        messages.error(request, "Course code and name are required.")
        return _safe_redirect_back(request, "manage_courses")

    has_program = _has_column('COURSE', 'program_id')
    has_is_active = _has_column('COURSE', 'is_active')

    with connection.cursor() as c:
        try:
            c.execute("SELECT COUNT(*) FROM COURSE WHERE course_code=%s AND course_id<>%s", [code, course_id])
            if (c.fetchone() or (0,))[0] > 0:
                messages.error(request, "Course code already in use.")
                return _safe_redirect_back(request, "manage_courses")
        except Exception:
            pass

        if has_program:
            c.execute("""
                UPDATE COURSE SET course_code=%s, course_name=%s, program_id=%s
                WHERE course_id=%s
            """, [code, name, program_id, course_id])
        else:
            c.execute("UPDATE COURSE SET course_code=%s, course_name=%s WHERE course_id=%s", [code, name, course_id])

        if has_is_active:
            c.execute("UPDATE COURSE SET is_active=%s WHERE course_id=%s", [0 if status == "inactive" else 1, course_id])

    messages.success(request, "Course updated." if status == "active" else "Course updated and set to Inactive.")
    return _safe_redirect_back(request, "manage_courses")


def restore_course(request, course_id):
    if not _has_column('COURSE', 'is_active'):
        messages.error(request, "Restore not available on this database schema.")
        return _safe_redirect_back(request, "manage_courses")

    with connection.cursor() as c:
        c.execute("UPDATE COURSE SET is_active=1 WHERE course_id=%s", [course_id])

    messages.success(request, "Course restored.")
    return _safe_redirect_back(request, "manage_courses")


def hard_delete_course(request, course_id):
    with connection.cursor() as c:
        c.execute("SELECT COUNT(*) FROM COURSE WHERE course_id=%s", [course_id])
        if (c.fetchone() or (0,))[0] == 0:
            messages.info(request, "Course not found.")
            return _safe_redirect_back(request, "manage_courses")

        has_is_active = _has_column('COURSE', 'is_active')
        is_active = 1
        if has_is_active:
            c.execute("SELECT is_active FROM COURSE WHERE course_id=%s", [course_id])
            (is_active,) = c.fetchone() or (1,)

        # usage (to allow delete when unused OR inactive)
        has_at_is_active = _has_column('ASSIGNED_TEACHER', 'is_active')
        if has_at_is_active:
            c.execute("SELECT COUNT(*) FROM ASSIGNED_TEACHER WHERE course_id=%s AND is_active=1", [course_id])
        else:
            c.execute("SELECT COUNT(*) FROM ASSIGNED_TEACHER WHERE course_id=%s", [course_id])
        assigned_now = (c.fetchone() or (0,))[0]

        c.execute("""
            SELECT COUNT(*)
            FROM LAB_SCHEDULE s
            JOIN ASSIGNED_TEACHER at ON at.assigned_teacher_id = s.assigned_teacher_id
            WHERE at.course_id=%s AND s.date >= CAST(GETDATE() AS DATE)
        """, [course_id])
        future_sched = (c.fetchone() or (0,))[0]

        slip_refs = 0
        if _has_column('UTILIZATION_SLIP', 'course_id'):
            c.execute("SELECT COUNT(*) FROM UTILIZATION_SLIP WHERE course_id=%s", [course_id])
            slip_refs = (c.fetchone() or (0,))[0]

        unused = (assigned_now == 0 and future_sched == 0 and slip_refs == 0)

        if int(is_active) == 1 and not unused:
            messages.error(request, "Cannot delete: active course is currently in use.")
            return _safe_redirect_back(request, "manage_courses")

        try:
            c.execute("DELETE FROM COURSE WHERE course_id=%s", [course_id])
        except Exception as e:
            messages.error(request, f"Delete failed (references may exist): {e}")
            return _safe_redirect_back(request, "manage_courses")

    messages.success(request, "Course permanently deleted.")
    return _safe_redirect_back(request, "manage_courses")


# ---------------- SEMESTERS ----------------

@transaction.atomic
def add_semester(request):
    if request.method != "POST":
        return _safe_redirect_back(request, "manage_courses")

    term = (request.POST.get("term") or "").strip()
    sy = (request.POST.get("school_year") or "").strip()
    sd = _parse_date(request.POST.get("start_date"))
    ed = _parse_date(request.POST.get("end_date"))

    if not term or not sy or not sd or not ed:
        messages.error(request, "All fields are required.")
        return _safe_redirect_back(request, "manage_courses")
    if ed < sd:
        messages.error(request, "End date cannot be earlier than start date.")
        return _safe_redirect_back(request, "manage_courses")

    with connection.cursor() as c:
        if _has_column('SEMESTER', 'is_active'):
            c.execute("""
                INSERT INTO SEMESTER (term, school_year, start_date, end_date, is_active)
                VALUES (%s, %s, %s, %s, 1)
            """, [term, sy, sd, ed])
        else:
            c.execute("""
                INSERT INTO SEMESTER (term, school_year, start_date, end_date)
                VALUES (%s, %s, %s, %s)
            """, [term, sy, sd, ed])

    messages.success(request, "Semester created.")
    return _safe_redirect_back(request, "manage_courses")


@transaction.atomic
def edit_semester(request, semester_id):
    if request.method != "POST":
        return HttpResponse("Invalid access", status=400)

    term = (request.POST.get("term") or "").strip()
    sy = (request.POST.get("school_year") or "").strip()
    sd = _parse_date(request.POST.get("start_date"))
    ed = _parse_date(request.POST.get("end_date"))
    status = (request.POST.get("status") or "active").strip().lower()

    if not term or not sy or not sd or not ed:
        messages.error(request, "All fields are required.")
        return _safe_redirect_back(request, "manage_courses")
    if ed < sd:
        messages.error(request, "End date cannot be earlier than start date.")
        return _safe_redirect_back(request, "manage_courses")

    with connection.cursor() as c:
        c.execute("""
            UPDATE SEMESTER
               SET term=%s, school_year=%s, start_date=%s, end_date=%s
             WHERE semester_id=%s
        """, [term, sy, sd, ed, semester_id])

        if _has_column('SEMESTER', 'is_active'):
            c.execute("UPDATE SEMESTER SET is_active=%s WHERE semester_id=%s",
                      [0 if status == "inactive" else 1, semester_id])

    messages.success(request, "Semester updated." if status == "active" else "Semester updated and set to Inactive.")
    return _safe_redirect_back(request, "manage_courses")


def restore_semester(request, semester_id):
    if not _has_column('SEMESTER', 'is_active'):
        messages.error(request, "Restore not available on this database schema.")
        return _safe_redirect_back(request, "manage_courses")

    with connection.cursor() as c:
        c.execute("UPDATE SEMESTER SET is_active=1 WHERE semester_id=%s", [semester_id])

    messages.success(request, "Semester restored.")
    return _safe_redirect_back(request, "manage_courses")


def hard_delete_semester(request, semester_id):
    with connection.cursor() as c:
        c.execute("SELECT COUNT(*) FROM SEMESTER WHERE semester_id=%s", [semester_id])
        if (c.fetchone() or (0,))[0] == 0:
            messages.info(request, "Semester not found.")
            return _safe_redirect_back(request, "manage_courses")

        is_active = 1
        if _has_column('SEMESTER', 'is_active'):
            c.execute("SELECT is_active FROM SEMESTER WHERE semester_id=%s", [semester_id])
            (is_active,) = c.fetchone() or (1,)

        c.execute("SELECT COUNT(*) FROM ASSIGNED_TEACHER WHERE semester_id=%s", [semester_id])
        in_assign = (c.fetchone() or (0,))[0]
        in_util = 0
        try:
            c.execute("SELECT COUNT(*) FROM UTILIZATION_SLIP WHERE semester_id=%s", [semester_id])
            in_util = (c.fetchone() or (0,))[0]
        except Exception:
            pass

        unused = (in_assign == 0 and in_util == 0)
        if int(is_active) == 1 and not unused:
            messages.error(request, "Cannot delete: active semester is currently in use.")
            return _safe_redirect_back(request, "manage_courses")

        try:
            c.execute("DELETE FROM SEMESTER WHERE semester_id=%s", [semester_id])
        except Exception as e:
            messages.error(request, f"Delete failed (references may exist): {e}")
            return _safe_redirect_back(request, "manage_courses")

    messages.success(request, "Semester permanently deleted.")
    return _safe_redirect_back(request, "manage_courses")


# ---------------- PROGRAMS ----------------

@transaction.atomic
def add_program(request):
    if request.method != "POST":
        return _safe_redirect_back(request, "manage_courses")

    if not _has_table('PROGRAM'):
        messages.error(request, "PROGRAM table is missing.")
        return _safe_redirect_back(request, "manage_courses")

    code = (request.POST.get("program_code") or "").strip().upper()
    name = (request.POST.get("program_name") or "").strip()

    if not code or not name:
        messages.error(request, "Program code and name are required.")
        return _safe_redirect_back(request, "manage_courses")

    with connection.cursor() as c:
        try:
            c.execute("SELECT COUNT(*) FROM PROGRAM WHERE program_code=%s", [code])
            if (c.fetchone() or (0,))[0] > 0:
                messages.error(request, "Program code already exists.")
                return _safe_redirect_back(request, "manage_courses")
        except Exception:
            pass

        if _has_column('PROGRAM', 'is_active'):
            c.execute("""
                INSERT INTO PROGRAM (program_code, program_name, is_active)
                VALUES (%s, %s, 1)
            """, [code, name])
        else:
            c.execute("""
                INSERT INTO PROGRAM (program_code, program_name)
                VALUES (%s, %s)
            """, [code, name])

    messages.success(request, "Program created.")
    return _safe_redirect_back(request, "manage_courses")


@transaction.atomic
def edit_program(request, program_id):
    if request.method != "POST":
        return HttpResponse("Invalid access", status=400)

    if not _has_table('PROGRAM'):
        return HttpResponseBadRequest("PROGRAM table is missing.")

    code = (request.POST.get("program_code") or "").strip().upper()
    name = (request.POST.get("program_name") or "").strip()
    status = (request.POST.get("status") or "active").strip().lower()

    if not code or not name:
        messages.error(request, "Program code and name are required.")
        return _safe_redirect_back(request, "manage_courses")

    has_is_active = _has_column('PROGRAM', 'is_active')

    with connection.cursor() as c:
        try:
            c.execute("SELECT program_id FROM PROGRAM WHERE program_code=%s", [code])
            row = c.fetchone()
            if row and int(row[0]) != int(program_id):
                messages.error(request, "Program code already in use.")
                return _safe_redirect_back(request, "manage_courses")
        except Exception:
            pass

        c.execute("""
            UPDATE PROGRAM
               SET program_code=%s, program_name=%s
             WHERE program_id=%s
        """, [code, name, program_id])

        if has_is_active:
            c.execute("UPDATE PROGRAM SET is_active=%s WHERE program_id=%s",
                      [0 if status == "inactive" else 1, program_id])

    messages.success(request, "Program updated." if status == "active" else "Program updated and set to Inactive.")
    return _safe_redirect_back(request, "manage_courses")


def restore_program(request, program_id):
    if not _has_column('PROGRAM', 'is_active'):
        messages.error(request, "Restore not available on this database schema.")
        return _safe_redirect_back(request, "manage_courses")

    with connection.cursor() as c:
        c.execute("UPDATE PROGRAM SET is_active=1 WHERE program_id=%s", [program_id])

    messages.success(request, "Program restored.")
    return _safe_redirect_back(request, "manage_courses")


def hard_delete_program(request, program_id):
    if not _has_table('PROGRAM'):
        return HttpResponseBadRequest("PROGRAM table is missing.")

    with connection.cursor() as c:
        c.execute("SELECT COUNT(*) FROM PROGRAM WHERE program_id=%s", [program_id])
        if (c.fetchone() or (0,))[0] == 0:
            messages.info(request, "Program not found.")
            return _safe_redirect_back(request, "manage_courses")

        is_active = 1
        if _has_column('PROGRAM', 'is_active'):
            c.execute("SELECT is_active FROM PROGRAM WHERE program_id=%s", [program_id])
            (is_active,) = c.fetchone() or (1,)

        ref_cnt = 0
        if _has_column('COURSE', 'program_id'):
            c.execute("SELECT COUNT(*) FROM COURSE WHERE program_id=%s", [program_id])
            ref_cnt = (c.fetchone() or (0,))[0]

        unused = (ref_cnt == 0)
        # allow delete when unused OR inactive
        if int(is_active) == 1 and not unused:
            messages.error(request, "Cannot delete: active program still used by courses.")
            return _safe_redirect_back(request, "manage_courses")

        try:
            c.execute("DELETE FROM PROGRAM WHERE program_id=%s", [program_id])
        except Exception as e:
            messages.error(request, f"Delete failed (courses may still reference it): {e}")
            return _safe_redirect_back(request, "manage_courses")

    messages.success(request, "Program permanently deleted.")
    return _safe_redirect_back(request, "manage_courses")
# MANAGE STUDENTS

def _has_column(table_name: str, column_name: str) -> bool:
    """SQL Server column-existence check using INFORMATION_SCHEMA."""
    with connection.cursor() as c:
        c.execute("""
            SELECT 1
            FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_NAME = %s AND COLUMN_NAME = %s
        """, [table_name, column_name])
        return c.fetchone() is not None




def _student_blockers(student_id: int) -> dict:
    """Counts child rows that block permanent delete."""
    blockers = {"attendance": 0}
    with connection.cursor() as c:
        if _has_column("COMPUTER_LAB_ATTENDANCE", "student_id"):
            c.execute("SELECT COUNT(*) FROM COMPUTER_LAB_ATTENDANCE WHERE student_id=%s", [student_id])
            blockers["attendance"] = int((c.fetchone() or (0,))[0])
    return blockers


def get_admin_header_context(user_id):
    """Minimal header context. Keep your own implementation if you already have it."""
    return {
        "current_page": "Manage Student",
        "teacher_name": "Admin",
        "teacher_profile_image": None,
        "notifications": [],
        "unread_count": 0,
    }


# ---------- list ----------
def manage_students(request):
    if not request.session.get('user_id') or request.session.get('role') != 'admin':
        return redirect('login')

    admin_id = request.session.get('user_id')
    context = build_user_header_context(admin_id)   # ← header here

    show_archived = request.GET.get('archived') == '1'
    q = (request.GET.get('q') or '').strip()

    def _safe_join(*parts):
        return " ".join(p.strip() for p in parts if p).strip()

    # 🔹 NEW: pull skipped upload info from session (if any)
    upload_skipped = request.session.pop('upload_skipped', None)


    params = [1 if show_archived else 0]
    where_extra = ""
    if q:
        like = f"%{q}%"
        where_extra = """
            AND (
                first_name LIKE %s OR
                middle_name LIKE %s OR
                last_name LIKE %s OR
                (first_name + ' ' + ISNULL(middle_name,'') + ' ' + last_name) LIKE %s OR
                stud_num LIKE %s OR
                email LIKE %s OR
                rfid_sticker_id LIKE %s
            )
        """
        params.extend([like, like, like, like, like, like, like])

    with connection.cursor() as cursor:
        cursor.execute(f"""
            SELECT student_id, first_name, middle_name, last_name, stud_num, email, rfid_sticker_id, ISNULL(is_archived,0)
            FROM STUDENTS
            WHERE ISNULL(is_archived,0) = %s
            {where_extra}
            ORDER BY last_name, first_name, middle_name
        """, params)
        rows = cursor.fetchall()

    students = []
    attendance_counts = {}
    if rows:
        student_ids = [r[0] for r in rows]
        with connection.cursor() as c:
            c.execute(f"""
                SELECT student_id, COUNT(*)
                FROM COMPUTER_LAB_ATTENDANCE
                WHERE student_id IN ({",".join(["%s"]*len(student_ids))})
                GROUP BY student_id
            """, student_ids)
            attendance_counts = dict(c.fetchall())

        for sid, fn, mn, ln, studnum, email, rfid, _arch in rows:
            full_name = _safe_join(fn, mn, ln)
            att = int(attendance_counts.get(sid, 0) or 0)
            can_hard_delete = (att == 0)
            delete_reason = None if can_hard_delete else f"Lab Attendance ({att})"
            students.append({
                "id": sid,
                "name": full_name or "—",
                "student_number": studnum,
                "email": email,
                "rfid": rfid or "N/A",
                "can_hard_delete": can_hard_delete,
                "delete_reason": delete_reason,
            })

    context.update({
        "students": students,
        "show_archived": show_archived,
        "q": q,
        "current_page": "Manage Student",
        "upload_skipped": upload_skipped,   # 🔹 NEW: pass to template
    })
    return render(request, "admin/manage_students.html", context)



# ---------- create (plain POST) ----------
@require_POST
def add_student(request):
    """Plain POST → redirect with messages; sends simple student/RFID email."""
    if not request.session.get('user_id') or request.session.get('role') != 'admin':
        return redirect('login')

    first_name = (request.POST.get('first_name') or '').strip()
    middle_name= (request.POST.get('middle_name') or '').strip()
    last_name  = (request.POST.get('last_name') or '').strip()
    email      = (request.POST.get('email') or '').strip()
    stud_num   = (request.POST.get('stud_num') or '').strip()

    if not first_name or not last_name or not email or not stud_num:
        messages.error(request, "First/Last name, email, and student number are required.")
        return redirect('manage_students')

    # normalize
    email_norm    = email.lower().strip()
    stud_num_norm = stud_num.strip()

    try:
        with connection.cursor() as cursor:
            # strict duplicate check (case-insensitive email)
            cursor.execute("""
                SELECT COUNT(*) 
                  FROM STUDENTS 
                 WHERE LOWER(email) = %s OR stud_num = %s
            """, [email_norm, stud_num_norm])
            if (cursor.fetchone() or [0])[0] > 0:
                messages.error(request, "Email or Student Number already exists.")
                return redirect('manage_students')

            cursor.execute("""
                INSERT INTO STUDENTS 
                    (first_name, middle_name, last_name, email, stud_num, rfid_sticker_id, role, is_archived)
                VALUES (%s, %s, %s, %s, %s, NULL, 'student', 0)
            """, [first_name, middle_name, last_name, email_norm, stud_num_norm])

    except IntegrityError as e:
        messages.error(request, f"Could not add student: {e}")
        return redirect('manage_students')

    # send simple "account created + go to admin for RFID" email
    try:
        full_name = " ".join([first_name, middle_name, last_name]).strip() or stud_num_norm
        send_student_rfid_account_email(
            to_email=email_norm,
            full_name=full_name,
            stud_num=stud_num_norm,
        )
        messages.success(request, "Student successfully added. Account email sent.")
    except Exception:
        messages.warning(request, "Student added, but email could not be sent.")

    return redirect('manage_students')


# ---------- edit (modal loads JSON; submit is plain POST) ----------

@require_GET
def get_student_json(request, student_id):
    if not request.session.get('user_id') or request.session.get('role') != 'admin':
        return JsonResponse({"error": "Unauthorized"}, status=401)

    with connection.cursor() as c:
        c.execute("""
            SELECT first_name, middle_name, last_name, email,
                   stud_num, ISNULL(is_archived,0) AS is_archived
            FROM STUDENTS WHERE student_id=%s
        """, [student_id])
        row = c.fetchone()

    if not row:
        return JsonResponse({"error": "Not found"}, status=404)

    status = "inactive" if int(row[5]) == 1 else "active"
    return JsonResponse({
        "first_name": row[0] or "",
        "middle_name": row[1] or "",
        "last_name": row[2] or "",
        "email": row[3] or "",
        "stud_num": row[4] or "",
        "status": status,
    })


@require_POST
def save_student(request, student_id):
    """Plain POST → redirect with messages (no JSON)."""
    if not request.session.get('user_id') or request.session.get('role') != 'admin':
        messages.error(request, "Unauthorized.")
        return redirect('manage_students')

    first_name = (request.POST.get('first_name') or '').strip()
    middle_name= (request.POST.get('middle_name') or '').strip()
    last_name  = (request.POST.get('last_name') or '').strip()
    email      = (request.POST.get('email') or '').strip()
    stud_num   = (request.POST.get('stud_num') or '').strip()
    status     = (request.POST.get('status') or 'active').strip().lower()

    if not first_name or not last_name or not email or not stud_num:
        messages.error(request, "First, last, email, and student number are required.")
        return redirect('manage_students')

    # 🔹 normalize
    email_norm    = email.lower()
    stud_num_norm = stud_num.strip()

    with connection.cursor() as c:
        c.execute("SELECT ISNULL(is_archived,0) FROM STUDENTS WHERE student_id=%s", [student_id])
        row = c.fetchone()
    if not row:
        messages.error(request, "Student not found.")
        return redirect('manage_students')
    current_archived = int(row[0])

    # duplicates (exclude self)
    with connection.cursor() as c:
        c.execute("""
            SELECT COUNT(*) 
              FROM STUDENTS 
             WHERE LOWER(email) = %s AND student_id<>%s
        """, [email_norm, student_id])
        if (c.fetchone() or [0])[0] > 0:
            messages.error(request, "Email already in use by another student.")
            return redirect('manage_students')

        c.execute("""
            SELECT COUNT(*) 
              FROM STUDENTS 
             WHERE stud_num=%s AND student_id<>%s
        """, [stud_num_norm, student_id])
        if (c.fetchone() or [0])[0] > 0:
            messages.error(request, "Student Number already in use by another student.")
            return redirect('manage_students')

    # update fields
    with connection.cursor() as c:
        c.execute("""
            UPDATE STUDENTS
               SET first_name=%s, middle_name=%s, last_name=%s,
                   email=%s, stud_num=%s
             WHERE student_id=%s
        """, [first_name, middle_name, last_name, email_norm, stud_num_norm, student_id])

    # status transitions
    if status == "inactive" and current_archived == 0:
        with connection.cursor() as c:
            c.execute("""
                UPDATE STUDENTS SET is_archived=1, rfid_sticker_id=NULL
                WHERE student_id=%s
            """, [student_id])
        messages.success(request, "Student updated and set to Inactive.")
    elif status == "active" and current_archived == 1:
        with connection.cursor() as c:
            c.execute("UPDATE STUDENTS SET is_archived=0 WHERE student_id=%s", [student_id])
        messages.success(request, "Student updated and reactivated.")
    else:
        messages.success(request, "Student updated.")

    return redirect('manage_students')



# ---------- Archive / Restore / Hard delete ----------

@require_POST
def delete_student(request, student_id):
    """Archive (soft delete)."""
    if not request.session.get('user_id') or request.session.get('role') != 'admin':
        return redirect('login')

    admin_id = request.session.get('user_id')

    with connection.cursor() as cursor:
        cursor.execute("""
            UPDATE STUDENTS SET is_archived = 1, rfid_sticker_id=NULL
             WHERE student_id = %s AND ISNULL(is_archived,0)=0
        """, [student_id])
        archived = cursor.rowcount or 0

        if archived:
            cursor.execute("""
                INSERT INTO ARCHIVE_LOG (user_type, user_id, action, performed_by)
                VALUES ('student', %s, 'archived', %s)
            """, [student_id, admin_id])

    if archived:
        messages.success(request, "Student archived.")
    else:
        messages.info(request, "Student already archived or not found.")
    return redirect('manage_students')


@require_POST
def restore_student(request, student_id):
    if not request.session.get('user_id') or request.session.get('role') != 'admin':
        return redirect('login')

    admin_id = request.session.get('user_id')
    with connection.cursor() as cursor:
        cursor.execute("""
            UPDATE STUDENTS SET is_archived = 0
             WHERE student_id = %s AND ISNULL(is_archived,0)=1
        """, [student_id])
        restored = cursor.rowcount or 0

        if restored:
            cursor.execute("""
                INSERT INTO ARCHIVE_LOG (user_type, user_id, action, performed_by)
                VALUES ('student', %s, 'restored', %s)
            """, [student_id, admin_id])

    if restored:
        messages.success(request, "Student restored.")
    else:
        messages.info(request, "Student was not archived or doesn’t exist.")
    return redirect('manage_students')


@require_POST
def hard_delete_student(request, student_id):
    """Permanent delete when no blockers exist."""
    if not request.session.get('user_id') or request.session.get('role') != 'admin':
        return redirect('login')

    with connection.cursor() as c:
        c.execute("SELECT COUNT(*) FROM STUDENTS WHERE student_id=%s", [student_id])
        if (c.fetchone() or (0,))[0] == 0:
            messages.info(request, "Student does not exist.")
            return redirect('manage_students')

    blockers = _student_blockers(student_id)
    blocking_keys = [k for k, v in blockers.items() if v and v > 0]
    if blocking_keys:
        labels = {"attendance": "Lab Attendance"}
        details = ", ".join(f"{labels[k]} ({blockers[k]})" for k in blocking_keys if k in labels)
        messages.error(request, f"Cannot delete: still referenced by {details}.")
        return redirect('manage_students')

    try:
        with transaction.atomic():
            with connection.cursor() as c:
                c.execute("DELETE FROM STUDENTS WHERE student_id=%s", [student_id])
                deleted = c.rowcount or 0
    except Exception as e:
        messages.error(request, f"Hard delete failed: {e}")
        return redirect('manage_students')

    if deleted:
        messages.success(request, "Student permanently deleted.")
    else:
        messages.info(request, "Student was already removed or does not exist.")
    return redirect('manage_students')


# ---------- RFID (assign-once) ----------

@csrf_exempt
def assign_rfid(request):
    """
    Assign RFID only if the student currently has no RFID.
    If already assigned, do NOT overwrite (return 'exists').
    """
    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "Invalid request."})

    student_id = request.POST.get("student_id")
    scanned_rfid = (request.POST.get("rfid") or '').strip()

    if not student_id or not scanned_rfid:
        return JsonResponse({"status": "error", "message": "Missing student or RFID."})

    with connection.cursor() as cursor:
        cursor.execute("SELECT rfid_sticker_id FROM STUDENTS WHERE student_id = %s", [student_id])
        row = cursor.fetchone()
        if not row:
            return JsonResponse({"status": "error", "message": "Student not found."})
        current_rfid = row[0]

        if current_rfid:
            return JsonResponse({"status": "exists", "message": "This student already has an RFID assigned."})

    if is_rfid_duplicate(scanned_rfid, exclude_student_id=student_id):
        return JsonResponse({"status": "duplicate", "message": "This RFID is already assigned to another entity."})

    with connection.cursor() as cursor:
        cursor.execute("""
            UPDATE STUDENTS
               SET rfid_sticker_id = %s
             WHERE student_id = %s
               AND (rfid_sticker_id IS NULL OR rfid_sticker_id = '')
        """, [scanned_rfid, student_id])

    return JsonResponse({"status": "success", "message": "RFID assigned successfully."})


# ---------- Bulk import ----------
def upload_excel(request):
    """
    Bulk import WITHOUT passwords.
    Expected columns (in order): first_name, middle_name, last_name, email, stud_num
    Sends simple student/RFID email for each successfully created student.
    """
    if not request.session.get('user_id') or request.session.get('role') != 'admin':
        return redirect('login')

    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"], "Invalid request method.")

    files = request.FILES.getlist('excel_files')
    if not files:
        return HttpResponse("No file uploaded.", status=400)

    total_created = 0
    total_skipped_count = 0
    email_sent_count = 0
    email_failed_count = 0
    skipped_rows = []  # collect detailed skipped info

    for f in files:
        fname = f.name.lower()
        if not fname.endswith('.xlsx'):
            total_skipped_count += 1
            continue

        try:
            wb = openpyxl.load_workbook(f, data_only=True)
        except Exception:
            total_skipped_count += 1
            continue

        sheet = wb.active

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row:
                total_skipped_count += 1
                skipped_rows.append({
                    "row": row_idx,
                    "first_name": "",
                    "middle_name": "",
                    "last_name": "",
                    "email": "",
                    "stud_num": "",
                    "reason": "Empty row.",
                })
                continue

            def _cell(i):
                return "" if i >= len(row) or row[i] is None else str(row[i]).strip()

            first_name = _cell(0)
            middle_name= _cell(1)
            last_name  = _cell(2)
            email      = _cell(3)
            stud_num   = _cell(4)

            email_norm    = email.lower().strip() if email else ""
            stud_num_norm = stud_num.strip() if stud_num else ""

            # Required fields check
            if not first_name or not last_name or not email_norm or not stud_num_norm:
                total_skipped_count += 1
                skipped_rows.append({
                    "row": row_idx,
                    "first_name": first_name,
                    "middle_name": middle_name,
                    "last_name": last_name,
                    "email": email,
                    "stud_num": stud_num,
                    "reason": "Missing required fields (First/Last name, Email, or Student Number).",
                })
                continue

            try:
                with connection.cursor() as cursor:
                    # duplicate check (case-insensitive email)
                    cursor.execute(
                        "SELECT COUNT(*) FROM STUDENTS WHERE LOWER(email)=LOWER(%s) OR stud_num=%s",
                        [email_norm, stud_num_norm]
                    )
                    if (cursor.fetchone() or [0])[0] > 0:
                        total_skipped_count += 1
                        skipped_rows.append({
                            "row": row_idx,
                            "first_name": first_name,
                            "middle_name": middle_name,
                            "last_name": last_name,
                            "email": email,
                            "stud_num": stud_num,
                            "reason": "Duplicate email or student number.",
                        })
                        continue

                    # insert student
                    cursor.execute("""
                        INSERT INTO STUDENTS (first_name, middle_name, last_name, email, stud_num,
                                              rfid_sticker_id, role, is_archived)
                        VALUES (%s, %s, %s, %s, %s, NULL, 'student', 0)
                    """, [first_name, middle_name, last_name, email_norm, stud_num_norm])

                total_created += 1

                # send student RFID/account email
                full_name = " ".join([first_name, middle_name, last_name]).strip() or stud_num_norm
                try:
                    send_student_rfid_account_email(
                        to_email=email_norm,
                        full_name=full_name,
                        stud_num=stud_num_norm,
                    )
                    email_sent_count += 1
                except Exception:
                    email_failed_count += 1

            except Exception:
                total_skipped_count += 1
                skipped_rows.append({
                    "row": row_idx,
                    "first_name": first_name,
                    "middle_name": middle_name,
                    "last_name": last_name,
                    "email": email,
                    "stud_num": stud_num,
                    "reason": "Unexpected error while inserting this row.",
                })
                continue

    if total_created:
        messages.success(request, f"Imported {total_created} student(s).")

    if email_sent_count:
        messages.success(request, f"Sent {email_sent_count} account email(s).")

    if email_failed_count:
        messages.warning(request, f"Failed to send {email_failed_count} account email(s). Check email settings/logs.")

    if total_skipped_count:
        messages.warning(request, f"Skipped {total_skipped_count} row(s).")

    if skipped_rows:
        request.session['upload_skipped'] = skipped_rows

    return redirect('manage_students')

# FACULTY
# ----------------- helpers -----------------

def _has_column(table, column) -> bool:
    with connection.cursor() as c:
        c.execute("""
            SELECT 1
            FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_NAME=%s AND COLUMN_NAME=%s
        """, [table, column])
        return c.fetchone() is not None

def _admin_count(include_archived: bool = False) -> int:
    with connection.cursor() as c:
        if include_archived:
            c.execute("SELECT COUNT(*) FROM FACULTY WHERE LOWER(role)='admin'")
        else:
            c.execute("SELECT COUNT(*) FROM FACULTY WHERE LOWER(role)='admin' AND ISNULL(is_archived,0)=0")
        (n,) = c.fetchone() or (0,)
        return n

def _full_name(row):
    # row = (first_name, middle_name, last_name, extension_name)
    first, mid, last, ext = row
    mid_i = f" {mid[0]}." if mid else ""
    ext_s = f" {ext}" if ext else ""
    return f"{(first or '').strip()}{mid_i} {(last or '').strip()}{ext_s}".strip()

# ---------- blockers used for hard delete gating ----------

def _faculty_blockers(faculty_id: int) -> dict:
    blockers = {}
    with connection.cursor() as c:
        c.execute("""
            SELECT COUNT(*)
            FROM LAB_SCHEDULE
            WHERE reserved_to=%s
               OR assigned_teacher_id IN (
                    SELECT assigned_teacher_id FROM ASSIGNED_TEACHER WHERE faculty_id=%s
               )
        """, [faculty_id, faculty_id])
        blockers["lab_schedule"] = (c.fetchone() or (0,))[0]

        if _has_column('ASSIGNED_TEACHER', 'is_active'):
            c.execute("""
                SELECT COUNT(*) FROM ASSIGNED_TEACHER
                WHERE faculty_id=%s AND ISNULL(is_active,1)=1
            """, [faculty_id])
        else:
            c.execute("SELECT COUNT(*) FROM ASSIGNED_TEACHER WHERE faculty_id=%s", [faculty_id])
        blockers["assigned_teacher"] = (c.fetchone() or (0,))[0]

        c.execute("""
            SELECT COUNT(*) FROM UTILIZATION_SLIP
            WHERE requested_by=%s OR processed_by=%s
        """, [faculty_id, faculty_id])
        blockers["utilization_slip"] = (c.fetchone() or (0,))[0]

        c.execute("SELECT COUNT(*) FROM LABORATORIES WHERE faculty_id=%s", [faculty_id])
        blockers["laboratories"] = (c.fetchone() or (0,))[0]

        c.execute("""
            SELECT COUNT(*) FROM NOTIFICATIONS
            WHERE receiver_teacher_id=%s OR sender_teacher_id=%s
        """, [faculty_id, faculty_id])
        blockers["notifications"] = (c.fetchone() or (0,))[0]

    return blockers

# ---------- LIST (Manage Faculty) ----------
def manage_faculty(request):
    if not request.session.get('user_id') or request.session.get('role') != 'admin':
        return redirect('login')

    admin_id = request.session.get('user_id')
    context = build_user_header_context(admin_id)

    # 🔹 Skipped rows from last upload (if any)
    upload_skipped = request.session.pop('teacher_upload_skipped', None)

    show_archived = request.GET.get('archived') == '1'
    q = (request.GET.get('q') or '').strip()

    params = [1 if show_archived else 0]
    where_extra = ""
    if q:
        like = f"%{q}%"
        where_extra = """
            AND (
                f.first_name LIKE %s OR
                f.middle_name LIKE %s OR
                f.last_name LIKE %s OR
                (f.first_name + ' ' + ISNULL(f.middle_name,'') + ' ' + f.last_name) LIKE %s OR
                f.email LIKE %s OR
                f.rfid_sticker_id LIKE %s
            )
        """
        params.extend([like, like, like, like, like, like])

    only_one_admin = (_admin_count(include_archived=False) <= 1)
    if not show_archived and only_one_admin:
        where_extra += " AND NOT (LOWER(f.role)='admin')"

    # 🔒 ALWAYS HIDE MAIN ADMIN (faculty_id = 1) FROM THE TABLE
    where_extra += " AND f.faculty_id <> %s"
    params.append(1)

    with connection.cursor() as cursor:
        cursor.execute(f"""
            SELECT
                f.faculty_id,
                f.first_name,
                f.middle_name,
                f.last_name,
                f.extension_name,
                f.email,
                f.rfid_sticker_id,
                LOWER(ISNULL(f.role,'teacher')) AS role
            FROM FACULTY f
            WHERE ISNULL(f.is_archived,0) = %s
            {where_extra}
            ORDER BY (CASE WHEN role='admin' THEN 0 ELSE 1 END),
                     f.last_name, f.first_name
        """, params)
        base_rows = cursor.fetchall()

    if not base_rows:
        # build empty ongoing semesters list for the page
        with connection.cursor() as c:
            sem_where = []
            if _has_column('SEMESTER', 'start_date') and _has_column('SEMESTER', 'end_date'):
                sem_where.append("CAST(start_date AS DATE) <= CAST(GETDATE() AS DATE)")
                sem_where.append("CAST(end_date   AS DATE) >= CAST(GETDATE() AS DATE)")
            if _has_column('SEMESTER', 'is_active'):
                sem_where.append("is_active = 1")
            elif _has_column('SEMESTER', 'is_archived'):
                sem_where.append("ISNULL(is_archived,0) = 0")

            sem_sql = "SELECT semester_id, term, school_year FROM SEMESTER"
            if sem_where:
                sem_sql += " WHERE " + " AND ".join(sem_where)
            sem_sql += " ORDER BY start_date DESC, semester_id DESC"
            c.execute(sem_sql)
            sem_rows = c.fetchall()
            all_semesters = [
                (sid, f"{(term or '').strip()} {(sy or '').strip()}".strip())
                for (sid, term, sy) in sem_rows
            ]

        context.update({
            "current_page": "Manage Teacher",
            "teachers": [],
            "show_archived": show_archived,
            "q": q,
            "all_courses": [],
            "all_semesters": all_semesters,
            "only_one_admin": only_one_admin,
            "upload_skipped": upload_skipped,
        })
        return render(request, "admin/manage_faculty.html", context)

    faculty_ids = [r[0] for r in base_rows]

    # ---------- course assignments ----------
    assignments_map = defaultdict(list)
    with connection.cursor() as c2:
        try:
            c2.execute(f"""
                SELECT at.faculty_id,
                       at.assigned_teacher_id, c.course_name, s.term, s.school_year
                  FROM ASSIGNED_TEACHER at
            INNER JOIN COURSE c   ON c.course_id   = at.course_id
            INNER JOIN SEMESTER s ON s.semester_id = at.semester_id
                 WHERE at.faculty_id IN ({",".join(["%s"]*len(faculty_ids))})
                   AND ISNULL(at.is_active,1)=1
                 ORDER BY c.course_name
            """, faculty_ids)
        except Exception:
            c2.execute(f"""
                SELECT at.faculty_id,
                       at.assigned_teacher_id, c.course_name, s.term, s.school_year
                  FROM ASSIGNED_TEACHER at
            INNER JOIN COURSE c   ON c.course_id   = at.course_id
            INNER JOIN SEMESTER s ON s.semester_id = at.semester_id
                 WHERE at.faculty_id IN ({",".join(["%s"]*len(faculty_ids))})
                 ORDER BY c.course_name
            """, faculty_ids)
        for fid, atid, cname, term, sy in c2.fetchall():
            assignments_map[fid].append({
                "atid": atid,
                "name": cname,
                "sem": f"{(term or '').strip()} {(sy or '').strip()}".strip()
            })

    # ---------- blockers prefetch ----------
    blockers = {
        fid: {
            "lab_schedule": 0,
            "assigned_teacher": 0,
            "utilization_slip": 0,
            "laboratories": 0,
            "notifications": 0
        } for fid in faculty_ids
    }

    with connection.cursor() as c:
        c.execute(f"""
            SELECT reserved_to, COUNT(*)
              FROM LAB_SCHEDULE
             WHERE reserved_to IN ({",".join(["%s"]*len(faculty_ids))})
             GROUP BY reserved_to
        """, faculty_ids)
        for fid, cnt in c.fetchall():
            if fid in blockers:
                blockers[fid]["lab_schedule"] += cnt

        c.execute(f"""
            SELECT at.faculty_id, COUNT(*)
              FROM LAB_SCHEDULE s
        INNER JOIN ASSIGNED_TEACHER at ON at.assigned_teacher_id = s.assigned_teacher_id
             WHERE at.faculty_id IN ({",".join(["%s"]*len(faculty_ids))})
             GROUP BY at.faculty_id
        """, faculty_ids)
        for fid, cnt in c.fetchall():
            if fid in blockers:
                blockers[fid]["lab_schedule"] += cnt

        try:
            c.execute(f"""
                SELECT faculty_id, COUNT(*)
                  FROM ASSIGNED_TEACHER
                 WHERE faculty_id IN ({",".join(["%s"]*len(faculty_ids))}) AND ISNULL(is_active,1)=1
                 GROUP BY faculty_id
            """, faculty_ids)
        except Exception:
            c.execute(f"""
                SELECT faculty_id, COUNT(*)
                  FROM ASSIGNED_TEACHER
                 WHERE faculty_id IN ({",".join(["%s"]*len(faculty_ids))})
                 GROUP BY faculty_id
            """, faculty_ids)
        for fid, cnt in c.fetchall():
            if fid in blockers:
                blockers[fid]["assigned_teacher"] = cnt

        c.execute(f"""
            SELECT requested_by, COUNT(*)
              FROM UTILIZATION_SLIP
             WHERE requested_by IN ({",".join(["%s"]*len(faculty_ids))})
             GROUP BY requested_by
        """, faculty_ids)
        for fid, cnt in c.fetchall():
            if fid in blockers:
                blockers[fid]["utilization_slip"] += cnt

        c.execute(f"""
            SELECT processed_by, COUNT(*)
              FROM UTILIZATION_SLIP
             WHERE processed_by IN ({",".join(["%s"]*len(faculty_ids))})
             GROUP BY processed_by
        """, faculty_ids)
        for fid, cnt in c.fetchall():
            if fid in blockers:
                blockers[fid]["utilization_slip"] += cnt

        c.execute(f"""
            SELECT faculty_id, COUNT(*)
              FROM LABORATORIES
             WHERE faculty_id IN ({",".join(["%s"]*len(faculty_ids))})
             GROUP BY faculty_id
        """, faculty_ids)
        for fid, cnt in c.fetchall():
            if fid in blockers:
                blockers[fid]["laboratories"] = cnt

        c.execute(f"""
            SELECT receiver_teacher_id, COUNT(*)
              FROM NOTIFICATIONS
             WHERE receiver_teacher_id IN ({",".join(["%s"]*len(faculty_ids))})
             GROUP BY receiver_teacher_id
        """, faculty_ids)
        for fid, cnt in c.fetchall():
            if fid in blockers:
                blockers[fid]["notifications"] += cnt

        c.execute(f"""
            SELECT sender_teacher_id, COUNT(*)
              FROM NOTIFICATIONS
             WHERE sender_teacher_id IN ({",".join(["%s"]*len(faculty_ids))})
             GROUP BY sender_teacher_id
        """, faculty_ids)
        for fid, cnt in c.fetchall():
            if fid in blockers:
                blockers[fid]["notifications"] += cnt

    labels = {
        "lab_schedule": "Lab Schedules",
        "assigned_teacher": "Course Assignments",
        "utilization_slip": "Utilization Slips",
        "laboratories": "Laboratories",
        "notifications": "Notifications",
    }

    teachers = []
    current_admin_id = request.session.get('user_id')

    # NOTE: base_rows now has 8 columns
    for (fid, fn, mn, ln, ext, email, rfid, role) in base_rows:
        is_protected_admin = (role == 'admin' and only_one_admin)

        # also protect currently logged-in admin from delete
        if int(fid) == int(current_admin_id) and role == 'admin':
            is_protected_admin = True

        can_hard_delete = False
        delete_reason = None
        if is_protected_admin:
            if int(fid) == int(current_admin_id) and role == 'admin':
                delete_reason = "You cannot delete your own admin account."
            else:
                delete_reason = "Cannot delete the last active admin."
        else:
            b = blockers.get(fid, {})
            blocking_keys = [k for k, v in b.items() if v and v > 0]
            if not blocking_keys:
                can_hard_delete = True
            else:
                delete_reason = ", ".join(f"{labels[k]} ({b[k]})" for k in blocking_keys)

        teachers.append({
            "id": fid,
            "first_name": fn,
            "middle_name": mn,
            "last_name": ln,
            "ext_name": ext,
            "email": email,
            "rfid": rfid or "N/A",
            "role": role,
            "courses": assignments_map.get(fid, []),
            "protected_admin": is_protected_admin,
            "can_hard_delete": can_hard_delete,
            "delete_reason": delete_reason,
        })

    with connection.cursor() as c:
        # courses
        course_where = []
        if _has_column('COURSE', 'is_active'):
            course_where.append("is_active = 1")
        elif _has_column('COURSE', 'is_archived'):
            course_where.append("ISNULL(is_archived,0) = 0")

        courses_sql = "SELECT course_id, course_name FROM COURSE"
        if course_where:
            courses_sql += " WHERE " + " AND ".join(course_where)
        courses_sql += " ORDER BY course_name"
        c.execute(courses_sql)
        all_courses = c.fetchall()

        # only current/ongoing semesters
        sem_where = []
        if _has_column('SEMESTER', 'start_date') and _has_column('SEMESTER', 'end_date'):
            sem_where.append("CAST(start_date AS DATE) <= CAST(GETDATE() AS DATE)")
            sem_where.append("CAST(end_date   AS DATE) >= CAST(GETDATE() AS DATE)")
        if _has_column('SEMESTER', 'is_active'):
            sem_where.append("is_active = 1")
        elif _has_column('SEMESTER', 'is_archived'):
            sem_where.append("ISNULL(is_archived,0) = 0")

        sem_sql = "SELECT semester_id, term, school_year FROM SEMESTER"
        if sem_where:
            sem_sql += " WHERE " + " AND ".join(sem_where)
        sem_sql += " ORDER BY start_date DESC, semester_id DESC"
        c.execute(sem_sql)
        sem_rows = c.fetchall()
        all_semesters = [
            (sid, f"{(term or '').strip()} {(sy or '').strip()}".strip())
            for (sid, term, sy) in sem_rows
        ]

    context.update({
        "current_page": "Manage Teacher",
        "teachers": teachers,
        "show_archived": show_archived,
        "q": q,
        "all_courses": all_courses,
        "all_semesters": all_semesters,
        "only_one_admin": only_one_admin,
        "current_admin_id": current_admin_id,
        "upload_skipped": upload_skipped,
    })
    return render(request, "admin/manage_faculty.html", context)

@require_POST
def upload_teacher_excel(request):
    if not request.session.get('user_id') or request.session.get('role') != 'admin':
        return redirect('login')

    files = request.FILES.getlist('excel_files')
    if not files:
        messages.error(request, "Please select at least one .xlsx file.")
        return redirect('manage_faculty')

    # existing emails (case-insensitive)
    existing_emails = set()
    with connection.cursor() as c:
        c.execute("SELECT LOWER(email) FROM FACULTY WHERE email IS NOT NULL")
        for (em,) in c.fetchall():
            if em:
                existing_emails.add(em)

    skipped_rows = []
    created_total = 0
    new_emails = set()  # prevent duplicates inside same batch

    for f in files:
        try:
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            ws = wb.active
        except Exception:
            messages.error(request, f"Could not read file {f.name}. Make sure it is a valid .xlsx file.")
            continue

        # header row
        header_row = [
            (cell.value or "").strip().lower() if cell.value else ""
            for cell in ws[1]
        ]

        def col_index(candidates):
            for name in candidates:
                if name in header_row:
                    return header_row.index(name)
            return None

        idx_first  = col_index(["first name", "first_name", "firstname"])
        idx_middle = col_index(["middle name", "middle_name", "middlename"])
        idx_last   = col_index(["last name", "last_name", "lastname"])
        idx_ext    = col_index(["extension name", "extension_name", "extensionname", "ext name", "ext"])
        idx_email  = col_index(["email", "email address", "email_address"])
        idx_pass   = col_index(["password", "temp password", "initial password"])
        idx_role   = col_index(["role", "user role"])

        # Required columns
        missing = []
        if idx_first is None: missing.append("First Name")
        if idx_last  is None: missing.append("Last Name")
        if idx_email is None: missing.append("Email")
        if idx_pass  is None: missing.append("Password")

        if missing:
            messages.error(request, f"{f.name}: Missing required columns: {', '.join(missing)}.")
            continue

        for excel_row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):

            def cell(idx):
                if idx is None:
                    return ""
                if idx >= len(row):
                    return ""
                return row[idx] if row[idx] is not None else ""

            first        = str(cell(idx_first)).strip()
            middle       = str(cell(idx_middle)).strip()
            last         = str(cell(idx_last)).strip()
            ext_name     = str(cell(idx_ext)).strip() if idx_ext is not None else ""
            email_raw    = str(cell(idx_email)).strip()
            password_raw = str(cell(idx_pass)).strip()
            role_raw     = str(cell(idx_role)).strip().lower() if idx_role is not None else ""

            # skip empty rows
            if not (first or middle or last or ext_name or email_raw or password_raw or role_raw):
                continue

            # missing required fields
            if not first or not last or not email_raw or not password_raw:
                skipped_rows.append({
                    "row": excel_row_index,
                    "first_name": first,
                    "middle_name": middle,
                    "last_name": last,
                    "extension_name": ext_name,
                    "email": email_raw,
                    "role": role_raw or "teacher",
                    "reason": "Missing required fields (first name, last name, email, or password)."
                })
                continue

            # normalize email
            email_norm = email_raw.lower()
            if "@" not in email_norm:
                skipped_rows.append({
                    "row": excel_row_index,
                    "first_name": first,
                    "middle_name": middle,
                    "last_name": last,
                    "extension_name": ext_name,
                    "email": email_raw,
                    "role": role_raw or "teacher",
                    "reason": "Invalid email format."
                })
                continue

            # default role
            if not role_raw:
                role = "teacher"
            elif role_raw not in ("teacher", "admin"):
                role = "teacher"
            else:
                role = role_raw

            # prevent duplicate emails
            if email_norm in existing_emails or email_norm in new_emails:
                skipped_rows.append({
                    "row": excel_row_index,
                    "first_name": first,
                    "middle_name": middle,
                    "last_name": last,
                    "extension_name": ext_name,
                    "email": email_raw,
                    "role": role,
                    "reason": "Email already exists."
                })
                continue

            try:
                # insert faculty
                with connection.cursor() as c:
                    c.execute("""
                        INSERT INTO FACULTY
                            (first_name, middle_name, last_name, extension_name, email, password, role)
                        VALUES (%s, %s, %s, %s, %s, %s, %s)
                    """, [
                        first,
                        middle or None,
                        last,
                        ext_name or None,
                        email_norm,
                        make_password(password_raw),
                        role
                    ])

                created_total += 1
                existing_emails.add(email_norm)
                new_emails.add(email_norm)

                # 🔔 send account email with login link
                try:
                    full_name_str = " ".join([p for p in [first, middle, last, ext_name] if p]).strip()
                    send_account_email_simple(
                        to_email=email_norm,
                        full_name=full_name_str,
                        role_label="Admin" if role == "admin" else "Faculty",
                        username_label="Email",
                        username_value=email_norm,
                        raw_password=password_raw,
                    )
                except Exception:
                    # you can log this if you want
                    pass

            except Exception as ex:
                skipped_rows.append({
                    "row": excel_row_index,
                    "first_name": first,
                    "middle_name": middle,
                    "last_name": last,
                    "extension_name": ext_name,
                    "email": email_raw,
                    "role": role,
                    "reason": f"Database error: {ex}"
                })

    if created_total:
        messages.success(request, f"Imported {created_total} teacher(s) successfully.")

    if skipped_rows:
        messages.warning(request, f"{len(skipped_rows)} row(s) were skipped. You can review them below.")
        request.session['teacher_upload_skipped'] = skipped_rows

    return redirect('manage_faculty')


# ---------- ADD ----------
def add_teacher(request):
    if not request.session.get('user_id') or request.session.get('role') != 'admin':
        return redirect('login')

    first_name     = (request.POST.get('first_name') or "").strip()
    middle_name    = (request.POST.get('middle_name') or "").strip()
    last_name      = (request.POST.get('last_name') or "").strip()
    extension_name = (request.POST.get('extension_name') or "").strip()
    email          = (request.POST.get('email') or "").strip()
    raw_password   = (request.POST.get('password') or "").strip()
    role           = (request.POST.get('role') or 'teacher').lower()

    if not first_name or not last_name or not email or not raw_password:
        messages.error(request, "All fields are required.")
        return redirect('manage_faculty')

    email_norm = email.lower()

    with connection.cursor() as cursor:
        cursor.execute("SELECT COUNT(*) FROM FACULTY WHERE LOWER(email)=%s", [email_norm])
        if (cursor.fetchone() or [0])[0] > 0:
            messages.error(request, "Email is already in use.")
            return redirect('manage_faculty')

        cursor.execute("""
            INSERT INTO FACULTY (first_name, middle_name, last_name, extension_name, email, password, role)
            OUTPUT INSERTED.faculty_id
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, [
            first_name,
            middle_name,
            last_name,
            extension_name or None,
            email_norm,
            make_password(raw_password),
            role
        ])
        _new_id = cursor.fetchone()[0]

    try:
        full_name_str = " ".join([p for p in [first_name, middle_name, last_name, extension_name] if p]).strip()
        send_account_email_simple(
            to_email=email_norm,
            full_name=full_name_str,
            role_label="Admin" if role == "admin" else "Faculty",
            username_label="Email",
            username_value=email_norm,
            raw_password=raw_password,
        )
    except Exception:
        pass

    messages.success(request, "Teacher successfully added.")
    return redirect('manage_faculty')

# ---------- EDIT (JSON) ----------
@require_GET
def get_teacher_json(request, faculty_id):
    if not request.session.get('user_id') or request.session.get('role') != 'admin':
        return JsonResponse({"error": "Unauthorized"}, status=401)

    with connection.cursor() as c:
        c.execute("""
            SELECT first_name,
                   middle_name,
                   last_name,
                   email,
                   LOWER(ISNULL(role,'teacher')) AS role,
                   ISNULL(is_archived,0) AS is_archived,
                   extension_name
            FROM FACULTY
            WHERE faculty_id=%s
        """, [faculty_id])
        row = c.fetchone()

    if not row:
        return JsonResponse({"error": "Not found"}, status=404)

    status = "inactive" if int(row[5]) == 1 else "active"
    return JsonResponse({
        "first_name": row[0] or "",
        "middle_name": row[1] or "",
        "last_name": row[2] or "",
        "email": row[3] or "",
        "role": row[4] or "teacher",
        "status": status,
        "extension_name": row[6] or "",
    })

@require_POST
def save_teacher_json(request, faculty_id):
    if not request.session.get('user_id') or request.session.get('role') != 'admin':
        return JsonResponse({"error": "Unauthorized"}, status=401)

    first_name     = (request.POST.get('first_name') or "").strip()
    middle_name    = (request.POST.get('middle_name') or "").strip()
    last_name      = (request.POST.get('last_name') or "").strip()
    extension_name = (request.POST.get('extension_name') or "").strip()
    email          = (request.POST.get('email') or "").strip()
    role           = (request.POST.get('role') or "teacher").strip().lower()
    status         = (request.POST.get('status') or "active").strip().lower()
    new_password   = (request.POST.get('password') or "").strip()

    # 🔒 prevent deactivating (archiving) your own account
    viewer_id = request.session.get('user_id')
    if int(faculty_id) == int(viewer_id) and status == "inactive":
        return JsonResponse({"error": "You cannot deactivate (archive) your own account."}, status=400)

    if not first_name or not last_name or not email:
        return JsonResponse({"error": "First, last, and email are required."}, status=400)

    with connection.cursor() as c:
        c.execute("SELECT LOWER(ISNULL(role,'teacher')), ISNULL(is_archived,0) FROM FACULTY WHERE faculty_id=%s", [faculty_id])
        row = c.fetchone()
    if not row:
        return JsonResponse({"error": "Not found"}, status=404)
    current_role, current_archived = row[0], int(row[1])

    if (role != "admin" or status == "inactive"):
        if current_role == "admin" and _admin_count(include_archived=False) <= 1:
            return JsonResponse({"error": "You cannot demote or inactivate the last active admin."}, status=400)

    with connection.cursor() as c:
        if new_password:
            pw_hash = make_password(new_password)
            c.execute("""
                UPDATE FACULTY
                   SET first_name=%s,
                       middle_name=%s,
                       last_name=%s,
                       extension_name=%s,
                       email=%s,
                       role=%s,
                       password=%s
                 WHERE faculty_id=%s
            """, [
                first_name,
                middle_name,
                last_name,
                extension_name or None,
                email,
                role,
                pw_hash,
                faculty_id
            ])
        else:
            c.execute("""
                UPDATE FACULTY
                   SET first_name=%s,
                       middle_name=%s,
                       last_name=%s,
                       extension_name=%s,
                       email=%s,
                       role=%s
                 WHERE faculty_id=%s
            """, [
                first_name,
                middle_name,
                last_name,
                extension_name or None,
                email,
                role,
                faculty_id
            ])

    if status == "inactive" and current_archived == 0:
        with connection.cursor() as c:
            c.execute("UPDATE FACULTY SET is_archived=1, rfid_sticker_id=NULL WHERE faculty_id=%s", [faculty_id])
            c.execute("""
                UPDATE LAB_SCHEDULE
                   SET status='Cancelled'
                 WHERE reserved_to=%s
                    OR assigned_teacher_id IN (
                        SELECT assigned_teacher_id FROM ASSIGNED_TEACHER WHERE faculty_id=%s
                    )
            """, [faculty_id, faculty_id])
            if _has_column('ASSIGNED_TEACHER', 'is_active'):
                c.execute("UPDATE ASSIGNED_TEACHER SET is_active=0 WHERE faculty_id=%s", [faculty_id])
            else:
                c.execute("DELETE FROM ASSIGNED_TEACHER WHERE faculty_id=%s", [faculty_id])

    if status == "active" and current_archived == 1:
        with connection.cursor() as c:
            c.execute("UPDATE FACULTY SET is_archived=0 WHERE faculty_id=%s", [faculty_id])

    return JsonResponse({"ok": True, "message": "Saved successfully."})


# ---------- RFID ----------
@csrf_exempt
def assign_teacher_rfid(request):
    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "Invalid request."})

    faculty_id = request.POST.get("faculty_id")
    scanned_rfid = (request.POST.get("rfid") or "").strip()

    if not faculty_id or not scanned_rfid:
        return JsonResponse({"status": "error", "message": "Missing teacher or RFID."})

    # Current RFID of this faculty (if any)
    with connection.cursor() as c:
        c.execute("""
            SELECT rfid_sticker_id 
            FROM FACULTY 
            WHERE faculty_id = %s AND ISNULL(is_archived,0) = 0
        """, [faculty_id])
        row = c.fetchone()
        current_rfid = row[0] if row else None

    # Same card already assigned to this teacher → nothing to change
    if current_rfid == scanned_rfid:
        return JsonResponse({
            "status": "success",
            "message": "RFID already assigned to this teacher."
        })

    # Check if some OTHER user already uses this RFID
    if is_rfid_duplicate(scanned_rfid, exclude_faculty_id=faculty_id):
        with connection.cursor() as c:
            # Try to find a faculty owner first (other teacher)
            c.execute("""
                SELECT TOP 1 faculty_id, first_name, middle_name, last_name, extension_name
                FROM FACULTY
                WHERE rfid_sticker_id = %s
                  AND faculty_id <> %s
                  AND ISNULL(is_archived,0) = 0
            """, [scanned_rfid, faculty_id])
            owner = c.fetchone()

            if owner:
                owner_name = _full_name(owner[1:])  # assumes you already have _full_name()
                return JsonResponse({
                    "status": "in_use",
                    "message": f"RFID is already assigned to {owner_name}.",
                    "owner_type": "faculty",
                    "owner_id": owner[0],
                    "owner_name": owner_name,
                })

            # If no faculty, maybe a student owns it
            c.execute("""
                SELECT TOP 1 student_id, first_name, middle_name, last_name
                FROM STUDENTS
                WHERE rfid_sticker_id = %s
                  AND ISNULL(is_archived,0) = 0
            """, [scanned_rfid])
            s_owner = c.fetchone()
            if s_owner:
                s_name = _full_name(s_owner[1:])
                return JsonResponse({
                    "status": "in_use",
                    "message": f"RFID is already assigned to student {s_name}.",
                    "owner_type": "student",
                    "owner_id": s_owner[0],
                    "owner_name": s_name,
                })

        # Fallback – duplicate but we couldn’t resolve owner (should be rare)
        return JsonResponse({
            "status": "duplicate",
            "message": "This RFID is already assigned."
        })

    # Free RFID → assign it
    with connection.cursor() as c:
        c.execute("""
            UPDATE FACULTY
            SET rfid_sticker_id = %s
            WHERE faculty_id = %s AND ISNULL(is_archived,0) = 0
        """, [scanned_rfid, faculty_id])

    return JsonResponse({
        "status": "success",
        "message": "RFID assigned successfully."
    })


# ---------- Assign / Unassign course ----------

@require_GET
def assign_course_options(request, faculty_id):
    """
    Returns:
      - programs:   [program_id, program_code, program_name]
      - courses:    [course_id, program_id, label "<code> — <name>"]
      - semesters:  [semester_id, label]   # only ongoing
      - assigned_map: {"<semester_id>": [course_id, ...]}  # active rows of this teacher
    """
    if not request.session.get('user_id') or request.session.get('role') != 'admin':
        return JsonResponse({"error": "Unauthorized"}, status=401)

    with connection.cursor() as c:
        c.execute("""
            SELECT program_id, program_code, program_name
            FROM PROGRAM
            WHERE ISNULL(is_active,1) = 1
            ORDER BY program_code ASC
        """)
        prows = c.fetchall()
        programs = [[pid, pcode, pname] for (pid, pcode, pname) in prows]

        course_where = []
        if _has_column('COURSE', 'is_active'):
            course_where.append("is_active = 1")
        elif _has_column('COURSE', 'is_archived'):
            course_where.append("ISNULL(is_archived,0) = 0")

        courses_sql = """
            SELECT c.course_id, c.program_id, (c.course_code + ' — ' + c.course_name) AS label
            FROM COURSE c
        """
        if course_where:
            courses_sql += " WHERE " + " AND ".join(course_where)
        courses_sql += " ORDER BY c.course_code ASC"
        c.execute(courses_sql)
        crows = c.fetchall()
        courses = [[int(cid), (pid if pid is not None else None), lbl] for (cid, pid, lbl) in crows]

        # --- Only the current/ongoing semester(s) for the dropdown ---
        sem_where = []
        if _has_column('SEMESTER', 'start_date') and _has_column('SEMESTER', 'end_date'):
            sem_where.append("CAST(start_date AS DATE) <= CAST(GETDATE() AS DATE)")
            sem_where.append("CAST(end_date   AS DATE) >= CAST(GETDATE() AS DATE)")
        if _has_column('SEMESTER', 'is_active'):
            sem_where.append("is_active = 1")
        elif _has_column('SEMESTER', 'is_archived'):
            sem_where.append("ISNULL(is_archived,0) = 0")

        sem_sql = "SELECT semester_id, term, school_year FROM SEMESTER"
        if sem_where:
            sem_sql += " WHERE " + " AND ".join(sem_where)
        sem_sql += " ORDER BY start_date DESC, semester_id DESC"
        c.execute(sem_sql)
        srows = c.fetchall()
        semesters = [[int(sid), f"{(t or '').strip()} {(sy or '').strip()}".strip()] for (sid, t, sy) in srows]

        # Map of this teacher's ACTIVE assignments per semester (for exclusion)
        if _has_column('ASSIGNED_TEACHER', 'is_active'):
            c.execute("""
                SELECT semester_id, course_id
                FROM ASSIGNED_TEACHER
                WHERE faculty_id = %s
                  AND ISNULL(is_active,1)=1
                  AND semester_id IS NOT NULL
                  AND course_id IS NOT NULL
            """, [faculty_id])
        else:
            c.execute("""
                SELECT semester_id, course_id
                FROM ASSIGNED_TEACHER
                WHERE faculty_id = %s
                  AND semester_id IS NOT NULL
                  AND course_id IS NOT NULL
            """, [faculty_id])

        assigned_map = {}
        for sid, cid in c.fetchall():
            if sid is None or cid is None:
                continue
            k = str(int(sid))
            assigned_map.setdefault(k, []).append(int(cid))

    return JsonResponse({
        "programs": programs,
        "courses": courses,
        "semesters": semesters,
        "assigned_map": assigned_map,
    })

@require_POST
def assign_course_save(request, faculty_id):
    """
    Create an active assignment. Validates:
      - course & semester are active
      - only ongoing semester (today within start/end)
      - if program_id provided: course belongs to that program
      - duplicate active assignment blocked
    """
    if not request.session.get('user_id') or request.session.get('role') != 'admin':
        return redirect('login')

    program_id = request.POST.get("program_id")
    course_id   = request.POST.get("course_id")
    semester_id = request.POST.get("semester_id")
    assigned_by = request.session.get("user_id")

    if not course_id or not semester_id:
        messages.error(request, "Please select course and semester.")
        return redirect("manage_faculty")

    with connection.cursor() as cursor:
        # course active? + get its program_id
        if _has_column('COURSE', 'is_active'):
            cursor.execute("SELECT program_id FROM COURSE WHERE course_id=%s AND is_active=1", [course_id])
        elif _has_column('COURSE', 'is_archived'):
            cursor.execute("SELECT program_id FROM COURSE WHERE course_id=%s AND ISNULL(is_archived,0)=0", [course_id])
        else:
            cursor.execute("SELECT program_id FROM COURSE WHERE course_id=%s", [course_id])
        row = cursor.fetchone()
        if not row:
            messages.error(request, "This course is archived/inactive or not found.")
            return redirect("manage_faculty")
        actual_program_id = row[0]

        if program_id and str(actual_program_id or "") != str(program_id or ""):
            messages.error(request, "Selected course does not belong to the chosen program.")
            return redirect("manage_faculty")

        # semester must be ongoing (today within range) and active if flags exist
        sem_checks = []
        if _has_column('SEMESTER', 'start_date') and _has_column('SEMESTER', 'end_date'):
            sem_checks.append("CAST(start_date AS DATE) <= CAST(GETDATE() AS DATE)")
            sem_checks.append("CAST(end_date   AS DATE) >= CAST(GETDATE() AS DATE)")
        if _has_column('SEMESTER', 'is_active'):
            sem_checks.append("is_active = 1")
        elif _has_column('SEMESTER', 'is_archived'):
            sem_checks.append("ISNULL(is_archived,0) = 0")

        where = "semester_id=%s"
        if sem_checks:
            where += " AND " + " AND ".join(sem_checks)

        cursor.execute(f"SELECT COUNT(*) FROM SEMESTER WHERE {where}", [semester_id])
        if (cursor.fetchone() or [0])[0] == 0:
            messages.error(request, "Selected semester is not currently active/ongoing.")
            return redirect("manage_faculty")

        # no duplicate active assignment
        try:
            cursor.execute("""
                SELECT COUNT(*)
                FROM ASSIGNED_TEACHER
                WHERE faculty_id=%s AND course_id=%s AND semester_id=%s AND ISNULL(is_active,1)=1
            """, [faculty_id, course_id, semester_id])
        except Exception:
            cursor.execute("""
                SELECT COUNT(*)
                FROM ASSIGNED_TEACHER
                WHERE faculty_id=%s AND course_id=%s AND semester_id=%s
            """, [faculty_id, course_id, semester_id])
        if (cursor.fetchone() or [0])[0] > 0:
            messages.error(request, "This teacher already has that course in the selected semester.")
            return redirect("manage_faculty")

        # insert
        try:
            cursor.execute("""
                INSERT INTO ASSIGNED_TEACHER (faculty_id, course_id, semester_id, assigned_by, is_active)
                VALUES (%s, %s, %s, %s, 1)
            """, [faculty_id, course_id, semester_id, assigned_by])
        except Exception:
            cursor.execute("""
                INSERT INTO ASSIGNED_TEACHER (faculty_id, course_id, semester_id, assigned_by)
                VALUES (%s, %s, %s, %s)
            """, [faculty_id, course_id, semester_id, assigned_by])

    messages.success(request, "Course assigned.")
    return redirect("manage_faculty")

@require_POST
def unassign_course(request, assigned_teacher_id):
    if not request.session.get('user_id') or request.session.get('role') != 'admin':
        return redirect('login')

    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT COUNT(*)
            FROM LAB_SCHEDULE
            WHERE assigned_teacher_id=%s AND date >= CAST(GETDATE() AS DATE)
        """, [assigned_teacher_id])
        (future_sched_count,) = cursor.fetchone() or (0,)
        if future_sched_count > 0:
            messages.error(request, "Cannot unassign: there are future schedules using this assignment.")
            return redirect(request.META.get("HTTP_REFERER") or "manage_faculty")

        try:
            cursor.execute("UPDATE ASSIGNED_TEACHER SET is_active=0 WHERE assigned_teacher_id=%s", [assigned_teacher_id])
        except Exception:
            cursor.execute("DELETE FROM ASSIGNED_TEACHER WHERE assigned_teacher_id=%s", [assigned_teacher_id])

    messages.success(request, "Course unassigned.")
    return redirect(request.META.get("HTTP_REFERER") or "manage_faculty")

# ---------- Archive / Restore / Hard delete (unchanged behavior) ----------
@require_POST
def delete_teacher(request, teacher_id):
    if not request.session.get('user_id') or request.session.get('role') != 'admin':
        return redirect('login')

    faculty_id = teacher_id
    admin_id = request.session.get('user_id')

    # 🔒 cannot archive yourself
    if int(faculty_id) == int(admin_id):
        messages.error(request, "You cannot archive your own admin account.")
        return redirect('manage_faculty')

    with connection.cursor() as c:
        c.execute("SELECT LOWER(ISNULL(role,'teacher')) FROM FACULTY WHERE faculty_id=%s", [faculty_id])
        role = (c.fetchone() or ["teacher"])[0]
    if role == "admin" and _admin_count(include_archived=False) <= 1:
        messages.error(request, "You cannot archive the last active admin.")
        return redirect('manage_faculty')

    with connection.cursor() as cursor:
        cursor.execute("""
            UPDATE FACULTY SET is_archived=1
            WHERE faculty_id=%s AND ISNULL(is_archived,0)=0
        """, [faculty_id])
        archived = cursor.rowcount or 0

        if archived:
            cursor.execute("""
                UPDATE LAB_SCHEDULE
                   SET status='Cancelled'
                 WHERE reserved_to=%s
                    OR assigned_teacher_id IN (
                        SELECT assigned_teacher_id FROM ASSIGNED_TEACHER WHERE faculty_id=%s
                    )
            """, [faculty_id, faculty_id])

            if _has_column('ASSIGNED_TEACHER', 'is_active'):
                cursor.execute("UPDATE ASSIGNED_TEACHER SET is_active=0 WHERE faculty_id=%s", [faculty_id])

            cursor.execute("""
                INSERT INTO ARCHIVE_LOG (user_type, user_id, action, performed_by)
                VALUES ('teacher', %s, 'archived', %s)
            """, [faculty_id, admin_id])

    if archived:
        messages.success(request, "Teacher archived. Related schedules were marked as Cancelled.")
    else:
        messages.info(request, "Teacher was already archived or does not exist.")
    return redirect('manage_faculty')


@require_POST
def restore_teacher(request, faculty_id):
    if not request.session.get('user_id') or request.session.get('role') != 'admin':
        return redirect('login')

    admin_id = request.session.get('user_id')

    with connection.cursor() as cursor:
        cursor.execute("""
            UPDATE FACULTY SET is_archived=0
            WHERE faculty_id=%s AND ISNULL(is_archived,0)=1
        """, [faculty_id])
        restored = cursor.rowcount or 0

        if restored:
            cursor.execute("""
                INSERT INTO ARCHIVE_LOG (user_type, user_id, action, performed_by)
                VALUES ('teacher', %s, 'restored', %s)
            """, [faculty_id, admin_id])

    if restored:
        messages.success(request, "Teacher restored.")
    else:
        messages.info(request, "Teacher was not archived or does not exist.")
    return redirect('manage_faculty')

@require_POST
def hard_delete_teacher(request, faculty_id):
    if not request.session.get('user_id') or request.session.get('role') != 'admin':
        return redirect('login')

    current_admin_id = request.session.get('user_id')

    # 🔒 cannot delete your own account
    if int(faculty_id) == int(current_admin_id):
        messages.error(request, "You cannot delete your own admin account.")
        return redirect('manage_faculty')

    with connection.cursor() as c:
        c.execute("SELECT LOWER(ISNULL(role,'teacher')), ISNULL(is_archived,0) FROM FACULTY WHERE faculty_id=%s", [faculty_id])
        row = c.fetchone()
    if not row:
        messages.info(request, "Teacher does not exist.")
        return redirect('manage_faculty')

    role, _is_archived = row[0], int(row[1])

    if role == "admin" and _admin_count(include_archived=False) <= 1:
        messages.error(request, "You cannot delete the last active admin.")
        return redirect('manage_faculty')

    blockers = _faculty_blockers(faculty_id)
    blocking_keys = [k for k, v in blockers.items() if v and v > 0]
    if blocking_keys:
        labels = {
            "lab_schedule": "Lab Schedules",
            "assigned_teacher": "Course Assignments",
            "utilization_slip": "Utilization Slips",
            "laboratories": "Laboratories",
            "notifications": "Notifications",
        }
        details = ", ".join(f"{labels[k]} ({blockers[k]})" for k in blocking_keys)
        messages.error(request, f"Cannot delete: still referenced by {details}.")
        return redirect('manage_faculty')

    try:
        with transaction.atomic():
            with connection.cursor() as c:
                c.execute("DELETE FROM ASSIGNED_TEACHER WHERE faculty_id = %s", [faculty_id])
                c.execute("DELETE FROM FACULTY WHERE faculty_id = %s", [faculty_id])
                deleted = c.rowcount or 0
    except Exception as e:
        messages.error(request, f"Hard delete failed: {e}")
        return redirect('manage_faculty')

    if deleted:
        messages.success(request, "Teacher permanently deleted.")
    else:
        messages.info(request, "Teacher was already removed or does not exist.")
    return redirect('manage_faculty')



# LABORATORIES

def _lab_blockers(lab_id: int) -> dict:
    """Return counts of referencing rows that should block hard delete."""
    blockers = {"lab_schedule": 0}
    with connection.cursor() as c:
        c.execute("SELECT COUNT(*) FROM LAB_SCHEDULE WHERE lab_id=%s", [lab_id])
        blockers["lab_schedule"] = (c.fetchone() or (0,))[0]
    return blockers


@require_GET
def manage_laboratories(request):
    if not request.session.get('user_id') or request.session.get('role') != 'admin':
        return redirect('login')

    admin_id = request.session.get('user_id')
    context = build_user_header_context(admin_id)

    show_archived = request.GET.get('archived') == '1'
    q = (request.GET.get('q') or '').strip()

    params = [1 if show_archived else 0]
    where_extra = ""
    if q:
        like = f"%{q}%"
        where_extra = """
            AND (
                CAST(l.lab_num AS NVARCHAR(20)) LIKE %s OR
                l.location LIKE %s OR
                ISNULL(f.first_name,'') LIKE %s OR
                ISNULL(f.last_name,'') LIKE %s OR
                ISNULL(l.rfid_reader_id,'') LIKE %s
            )
        """
        params.extend([like, like, like, like, like])

    with connection.cursor() as cursor:
        cursor.execute(f"""
            SELECT l.lab_id, l.lab_num, l.location, l.capacity, l.rfid_reader_id,
                   ISNULL(f.first_name,''), ISNULL(f.last_name,'')
              FROM LABORATORIES l
         LEFT JOIN FACULTY f
                ON l.faculty_id = f.faculty_id AND ISNULL(f.is_archived,0)=0
             WHERE ISNULL(l.is_archived,0) = %s
                   {where_extra}
          ORDER BY l.lab_num ASC
        """, params)
        rows = cursor.fetchall()

    labs = []
    for lab_id, lab_num, location, capacity, rfid_reader_id, fname, lname in rows:
        blockers = _lab_blockers(lab_id)
        blocking = [k for k, v in blockers.items() if v and v > 0]
        can_hard_delete = not bool(blocking)
        labels = {"lab_schedule": "Lab Schedules"}
        delete_reason = None if can_hard_delete else ", ".join(f"{labels[k]} ({blockers[k]})" for k in blocking)
        labs.append({
            "id": lab_id,
            "lab_num": lab_num,
            "location": location,
            "capacity": capacity,
            "rfid_reader_id": rfid_reader_id or "N/A",
            "in_charge": f"{(fname or '').strip()} {(lname or '').strip()}".strip() or "Unassigned",
            "can_hard_delete": can_hard_delete,
            "delete_reason": delete_reason,
        })

    with connection.cursor() as c:
        c.execute("""
            SELECT faculty_id, first_name, last_name
              FROM FACULTY
             WHERE ISNULL(is_archived,0)=0
             ORDER BY last_name, first_name
        """)
        faculty_list = c.fetchall()

    context.update({
        "laboratories": labs,
        "faculty_list": faculty_list,
        "current_page": "Manage Laboratories",
        "show_archived": show_archived,
        "q": q,
    })
    return render(request, "admin/manage_laboratories.html", context)


def insert_lab_schedule_if_assigned(cursor, lab_id, faculty_id):
    """Optional helper (safe no-op if not needed)."""
    if not faculty_id:
        return
    cursor.execute("""
        SELECT COUNT(*) FROM LAB_SCHEDULE
        WHERE lab_id=%s AND assigned_teacher_id IS NULL AND reserved_to=%s
    """, [lab_id, faculty_id])
    exists = (cursor.fetchone() or [0])[0]
    if exists == 0:
        cursor.execute("""
            INSERT INTO LAB_SCHEDULE (
                lab_id, reserved_to, status, date, start_time, end_time,
                student_year_and_section, number_of_students
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, [lab_id, faculty_id, 'Reserved', timezone.localdate(), time(0,0), time(0,0), '', 0])


@require_POST
def add_laboratory(request):
    if not request.session.get('user_id') or request.session.get('role') != 'admin':
        return redirect('login')

    lab_num = (request.POST.get('lab_num') or "").strip()
    location = (request.POST.get('location') or "").strip()
    capacity_raw = request.POST.get('capacity')
    faculty_id = request.POST.get('faculty_id') or None
    rfid_reader_id = (request.POST.get('rfid_reader_id') or "").strip() or None

    try:
        capacity = int(capacity_raw)
        if capacity < 1:
            raise ValueError
    except (ValueError, TypeError):
        messages.error(request, "Invalid capacity. Please enter a valid number.")
        return redirect('manage_laboratories')

    if rfid_reader_id and is_rfid_duplicate(rfid_reader_id):
        messages.error(request, "RFID Reader ID is already assigned to another record.")
        return redirect('manage_laboratories')

    with connection.cursor() as cursor:
        cursor.execute("""
            INSERT INTO LABORATORIES (lab_num, location, capacity, faculty_id, rfid_reader_id, is_archived)
            OUTPUT INSERTED.lab_id
            VALUES (%s, %s, %s, %s, %s, 0)
        """, [lab_num, location, capacity, faculty_id, rfid_reader_id])
        row = cursor.fetchone()
        if not row:
            messages.error(request, "Laboratory creation failed.")
            return redirect('manage_laboratories')

        lab_id = row[0]
        if faculty_id:
            insert_lab_schedule_if_assigned(cursor, lab_id, faculty_id)

    messages.success(request, "Laboratory successfully added.")
    return redirect('manage_laboratories')


# ---------- EDIT (AJAX JSON for modal) ----------

@require_GET
def get_laboratory_json(request, lab_id):
    if not request.session.get('user_id') or request.session.get('role') != 'admin':
        return JsonResponse({"error": "Unauthorized"}, status=401)

    with connection.cursor() as c:
        c.execute("""
            SELECT lab_num, location, capacity, faculty_id, rfid_reader_id, ISNULL(is_archived,0)
              FROM LABORATORIES WHERE lab_id=%s
        """, [lab_id])
        row = c.fetchone()

    if not row:
        return JsonResponse({"error": "Not found"}, status=404)

    status = "inactive" if int(row[5]) == 1 else "active"
    return JsonResponse({
        "lab_num": row[0],
        "location": row[1],
        "capacity": row[2],
        "faculty_id": row[3],
        "rfid_reader_id": row[4],
        "status": status,
    })


@require_POST
def save_laboratory_json(request, lab_id):
    """
    Update fields. If status=inactive:
      - LABORATORIES.is_archived = 1
      - LABORATORIES.rfid_reader_id = NULL
      - LAB_SCHEDULE(date >= today) -> status='Cancelled'
    If status=active:
      - LABORATORIES.is_archived = 0
    """
    if not request.session.get('user_id') or request.session.get('role') != 'admin':
        return JsonResponse({"error": "Unauthorized"}, status=401)

    with connection.cursor() as c:
        c.execute("SELECT ISNULL(is_archived,0) FROM LABORATORIES WHERE lab_id=%s", [lab_id])
        row = c.fetchone()
    if not row:
        return JsonResponse({"error": "Not found"}, status=404)
    current_archived = int(row[0])

    lab_num        = (request.POST.get('lab_num') or "").strip()
    location       = (request.POST.get('location') or "").strip()
    capacity_raw   = request.POST.get('capacity')
    faculty_id     = request.POST.get('faculty_id') or None
    rfid_reader_id = (request.POST.get('rfid_reader_id') or "").strip() or None
    status         = (request.POST.get('status') or 'active').strip().lower()

    try:
        capacity = int(capacity_raw)
        if capacity < 1:
            raise ValueError
    except (ValueError, TypeError):
        return JsonResponse({"error": "Invalid capacity."}, status=400)

    # If you have exclude support, enable this:
    # if rfid_reader_id and is_rfid_duplicate(rfid_reader_id, exclude_lab_id=lab_id):
    #     return JsonResponse({"error": "This RFID is already assigned to another entity."}, status=400)

    with connection.cursor() as c:
        c.execute("""
            UPDATE LABORATORIES
               SET lab_num=%s, location=%s, capacity=%s, faculty_id=%s, rfid_reader_id=%s
             WHERE lab_id=%s
        """, [lab_num, location, capacity, faculty_id, rfid_reader_id, lab_id])

    today = timezone.localdate()
    try:
        with connection.cursor() as c:
            if status == "inactive" and current_archived == 0:
                c.execute("UPDATE LABORATORIES SET is_archived=1, rfid_reader_id=NULL WHERE lab_id=%s", [lab_id])
                c.execute("UPDATE LAB_SCHEDULE SET status='Cancelled' WHERE lab_id=%s AND date >= %s", [lab_id, today])
            elif status == "active" and current_archived == 1:
                c.execute("UPDATE LABORATORIES SET is_archived=0 WHERE lab_id=%s", [lab_id])
    except Exception:
        return JsonResponse({"error": "Update failed."}, status=500)

    msg = "Saved successfully."
    if status == "inactive" and current_archived == 0:
        msg = "Saved and set to Inactive. Upcoming schedules were cancelled and RFID cleared."
    elif status == "active" and current_archived == 1:
        msg = "Saved and restored to Active."

    return JsonResponse({"ok": True, "message": msg})


# ---------- archive / restore / hard delete ----------

@require_POST
def archive_laboratory(request, lab_id):
    if not request.session.get('user_id') or request.session.get('role') != 'admin':
        return redirect('login')

    today = timezone.localdate()
    with connection.cursor() as c:
        c.execute("""
            UPDATE LABORATORIES
               SET is_archived=1, rfid_reader_id=NULL
             WHERE lab_id=%s AND ISNULL(is_archived,0)=0
        """, [lab_id])
        archived = c.rowcount or 0

        if archived:
            c.execute("""
                UPDATE LAB_SCHEDULE
                   SET status='Cancelled'
                 WHERE lab_id=%s AND date >= %s
            """, [lab_id, today])

    if archived:
        messages.success(request, "Laboratory archived. Upcoming schedules were cancelled and RFID cleared.")
    else:
        messages.info(request, "Laboratory was already archived or does not exist.")
    return redirect('manage_laboratories')


@require_POST
def restore_laboratory(request, lab_id):
    if not request.session.get('user_id') or request.session.get('role') != 'admin':
        return redirect('login')

    with connection.cursor() as c:
        c.execute("""
            UPDATE LABORATORIES
               SET is_archived=0
             WHERE lab_id=%s AND ISNULL(is_archived,0)=1
        """, [lab_id])
        restored = c.rowcount or 0

    if restored:
        messages.success(request, "Laboratory restored.")
    else:
        messages.info(request, "Laboratory was not archived or does not exist.")
    return redirect('manage_laboratories')


@require_POST
def hard_delete_laboratory(request, lab_id):
    if not request.session.get('user_id') or request.session.get('role') != 'admin':
        return redirect('login')

    blockers = _lab_blockers(lab_id)
    if blockers.get("lab_schedule", 0) > 0:
        messages.error(request, f"Cannot delete: still referenced by Lab Schedules ({blockers['lab_schedule']}).")
        return redirect('manage_laboratories')

    try:
        with transaction.atomic():
            with connection.cursor() as c:
                # c.execute("DELETE FROM LAB_SCHEDULE WHERE lab_id=%s", [lab_id])  # optional
                c.execute("DELETE FROM LABORATORIES WHERE lab_id=%s", [lab_id])
                deleted = c.rowcount or 0
    except Exception as e:
        messages.error(request, f"Hard delete failed: {e}")
        return redirect('manage_laboratories')

    if deleted:
        messages.success(request, "Laboratory permanently deleted.")
    else:
        messages.info(request, "Laboratory was already removed or does not exist.")
    return redirect('manage_laboratories')
# MANAGE SCHEDULE

def _monday_of(d: date) -> date:
    return d - timedelta(days=d.weekday())

def _sunday_of(d: date) -> date:
    return _monday_of(d) + timedelta(days=6)

def _db_day_of_week(py_weekday: int) -> int:
    """ OPERATING_TIME.day_of_week uses 0=Sun..6=Sat, Python uses 0=Mon..6=Sun """
    return (py_weekday + 1) % 7

def _safe_name(*parts) -> str:
    return " ".join(p.strip() for p in parts if p and isinstance(p, str)).strip() or "—"

def _short_name(first: str | None, last: str | None) -> str:
    """Return 'F. Last' when both present; otherwise whichever exists; fallback '—'."""
    f = (first or "").strip()
    l = (last or "").strip()
    if f and l:
        return f"{f[0]}. {l}"
    return l or f or "—"

def _fmt_hhmm_24(t: time | None) -> str:
    return t.strftime("%H:%M") if t else "—"

def _fmt_hhmm_12(t: time | None) -> str:
    # 12-hour with AM/PM (e.g. 01:05 PM); handle None safely.
    return t.strftime("%I:%M %p").lstrip("0") if t else "—"

# Alias used in some messages below
_fmt_hhmm = _fmt_hhmm_12


# Resolve course (fallback) – returns (course_code, course_name); we will only use code
def _resolve_course_for_slip_fallback(faculty_id: int, on_date: date):
    with connection.cursor() as c:
        c.execute("""
            SELECT TOP 1 c.course_code, c.course_name
              FROM ASSIGNED_TEACHER at
              JOIN SEMESTER s ON s.semester_id = at.semester_id
              JOIN COURSE   c ON c.course_id     = at.course_id
             WHERE at.faculty_id = %s
               AND %s BETWEEN s.start_date AND s.end_date
             ORDER BY s.start_date DESC
        """, [faculty_id, on_date])
        row = c.fetchone()
    return (row[0], row[1]) if row else (None, None)

# ---------------------------------------------------------
# Live: Active sessions
# ---------------------------------------------------------


# ---------------------------------------------------------
# Admin: Manage Schedule (page)
# ---------------------------------------------------------
@require_http_methods(["GET"])
def manage_schedule(request):
    if not request.session.get("user_id") or request.session.get("role") != "admin":
        return redirect("login")

    # Replace with your own context builder
    admin_id = request.session.get("user_id")
    context = build_user_header_context(admin_id)

    # Labs
    with connection.cursor() as cursor:
        cursor.execute("SELECT lab_id, lab_num FROM LABORATORIES ORDER BY lab_num;")
        labs = cursor.fetchall()

    if not labs:
        context.update({
            "current_page": "Lab Reservation",
            "reservations": [],
            "status_filter": "",
            "labs": [],
            "selected_lab": None,
            "week_range_display": "",
            "days": ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"],
            "pending_count": 0,
            "faculties": [],
            "faculty_courses_json": json.dumps({}),
            "config_json": json.dumps({}),
            "week_url": reverse("admin_week_data"),
            "base_week": timezone.localdate().isoformat(),
        })
        return render(request, "admin/manage_schedule.html", context)

    selected_lab = request.GET.get("lab_id") or str(labs[0][0])
    try:
        selected_lab_int = int(selected_lab)
    except (TypeError, ValueError):
        selected_lab_int = labs[0][0]

    # requests filter for modal
    allowed_status = {"Pending", "Approved", "Rejected"}
    status_filter = request.GET.get("status") or ""
    if status_filter not in allowed_status:
        status_filter = ""

    # faculties (active only)
    with connection.cursor() as c:
        c.execute("""
            SELECT faculty_id, (RTRIM(LTRIM(COALESCE(last_name,''))) + ', ' +
                                RTRIM(LTRIM(COALESCE(first_name,'')))) AS name
            FROM FACULTY
            WHERE ISNULL(is_archived,0)=0
            ORDER BY last_name, first_name;
        """)
        faculties = [(r[0], r[1]) for r in c.fetchall()]

    # faculty → courses mapping
    fac_courses = {}
    with connection.cursor() as c:
        c.execute("""
            SELECT at.faculty_id, c.course_id, COALESCE(c.course_code, c.course_name) AS label
            FROM ASSIGNED_TEACHER at
            JOIN COURSE c ON c.course_id = at.course_id
        """)
        for fid, cid, label in c.fetchall():
            fac_courses.setdefault(str(fid), []).append({"id": cid, "code": label})

    # Requests for modal (only course_code)
    where = ["u.lab_id = %s"]
    params = [selected_lab_int]
    if status_filter:
        where.append("u.status = %s")
        params.append(status_filter)
    where_sql = " AND ".join(where)

    with connection.cursor() as cursor:
        cursor.execute(f"""
            SELECT
                u.utilization_id,
                u.date,
                COALESCE(ls.start_time, u.start_time) AS st,
                COALESCE(ls.end_time,   u.end_time)   AS et,
                u.status, u.remarks,
                f.first_name, f.last_name,
                l.lab_num,
                p.first_name, p.last_name,
                u.student_year_and_section,
                c1.course_code,         -- schedule's course code
                c2.course_code,         -- slip's course code
                u.requested_by,
                u.schedule_id
            FROM UTILIZATION_SLIP u
            JOIN FACULTY f ON u.requested_by = f.faculty_id AND ISNULL(f.is_archived,0) = 0
            LEFT JOIN LABORATORIES l   ON u.lab_id = l.lab_id
            LEFT JOIN LAB_SCHEDULE ls  ON u.schedule_id = ls.schedule_id
            LEFT JOIN ASSIGNED_TEACHER at1 ON ls.assigned_teacher_id = at1.assigned_teacher_id
            LEFT JOIN COURSE          c1   ON at1.course_id = c1.course_id
            LEFT JOIN ASSIGNED_TEACHER at2 ON u.assigned_teacher_id = at2.assigned_teacher_id
            LEFT JOIN COURSE          c2   ON at2.course_id = c2.course_id
            LEFT JOIN FACULTY p ON u.processed_by = p.faculty_id
            WHERE {where_sql}
            ORDER BY u.date DESC, COALESCE(ls.start_time, u.start_time) DESC
        """, params)
        rows = cursor.fetchall()

    reservations = []
    fallback_cache = {}
    for r in rows:
        (util_id, d, st, et, status, remarks,
         r_fn, r_ln, lab_num,
         proc_fn, proc_ln,
         year_sec,
         sc_code, slip_code,
         requested_by,
         schedule_id) = r

        course_code = sc_code or slip_code
        if not course_code and d:
            key = (requested_by, d)
            if key not in fallback_cache:
                fallback_cache[key] = _resolve_course_for_slip_fallback(requested_by, d)
            fc_code, _fc_name = fallback_cache[key]
            course_code = course_code or fc_code

        reservations.append({
            "id": util_id,
            "name": _short_name(r_fn, r_ln),
            "lab": f"Lab {lab_num}" if lab_num else "—",
            "day_name": d.strftime("%A") if d else "—",
            "date": d,
            "start_time": _fmt_hhmm_12(st),   # 12h
            "end_time": _fmt_hhmm_12(et),     # 12h
            "status": status,
            "purpose": remarks or "",
            "processed_by_name": _short_name(proc_fn, proc_ln) if (proc_fn or proc_ln) else "—",
            "year_section": year_sec or "—",
            "course": course_code or "—",   # only code
        })

    # pending count
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT COUNT(*) FROM UTILIZATION_SLIP
             WHERE lab_id = %s AND status = 'Pending'
        """, [selected_lab_int])
        row = cursor.fetchone()
        pending_count = (row[0] if row else 0) or 0

    # Week window
    week_str = request.GET.get("week")
    try:
        base_date = dt.strptime(week_str, "%Y-%m-%d").date() if week_str else timezone.localdate()
    except Exception:
        base_date = timezone.localdate()
    start_w = _monday_of(base_date)
    end_w = _sunday_of(base_date)
    week_range_display = f"{start_w.strftime('%b %d')} - {end_w.strftime('%d, %Y')}"

    # Template context
    context.update({
        "current_page": "Lab Reservation",
        "labs": labs,
        "selected_lab": selected_lab_int,
        "week_range_display": week_range_display,
        "days": ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"],
        "reservations": reservations,
        "status_filter": status_filter,
        "pending_count": pending_count,
        "faculties": faculties,
        "faculty_courses_json": json.dumps(fac_courses),
        "config_json": json.dumps({
            "selected_lab": selected_lab_int,
            "base_week": start_w.isoformat(),
            "week_url": reverse("admin_week_data"),
        }),
        "week_url": reverse("admin_week_data"),
        "base_week": start_w.isoformat(),
    })

    return render(request, "admin/manage_schedule.html", context)

# ---------------------------------------------------------
# Week data JSON for grids
# ---------------------------------------------------------
@require_http_methods(["GET"])
def admin_week_data(request):
    lab_id = request.GET.get("lab_id")
    week_str = request.GET.get("week")
    include_requests = request.GET.get("include_requests") == "1"
    if not lab_id:
        return JsonResponse({"error": "lab_id required"}, status=400)

    try:
        base_date = dt.strptime(week_str, "%Y-%m-%d").date() if week_str else timezone.localdate()
    except Exception:
        base_date = timezone.localdate()

    start_w = _monday_of(base_date)
    end_w = _sunday_of(base_date)
    week_range_display = f"{start_w.strftime('%b %d')} - {end_w.strftime('%d, %Y')}"

    entries = defaultdict(lambda: defaultdict(list))
    req_entries = defaultdict(lambda: defaultdict(list))
    time_set = set()

    # --- SCHEDULES (hide only Cancelled/Rejected; keep Completed) ---
    with connection.cursor() as cur:
        cur.execute("""
            SELECT s.date, s.start_time, s.end_time,
                   f.first_name, f.last_name, f.faculty_id,
                   c.course_code,
                   s.student_year_and_section
            FROM LAB_SCHEDULE s
       LEFT JOIN ASSIGNED_TEACHER at ON s.assigned_teacher_id = at.assigned_teacher_id
       LEFT JOIN COURSE          c  ON at.course_id = c.course_id
       LEFT JOIN FACULTY         f  ON at.faculty_id = f.faculty_id AND ISNULL(f.is_archived,0)=0
           WHERE s.lab_id = %s
             AND s.date BETWEEN %s AND %s
             AND UPPER(LTRIM(RTRIM(ISNULL(s.status, '')))) NOT IN ('CANCELLED','REJECTED')  -- keep COMPLETED visible
           ORDER BY s.date, s.start_time
        """, [lab_id, start_w, end_w])
        for d, st, et, fn, ln, fid, course_code, section in cur.fetchall():
            dayname = d.strftime("%A")
            tl_12 = f"{_fmt_hhmm_12(st)} - {_fmt_hhmm_12(et)}"
            entries[dayname][tl_12].append({
                "teacher": _short_name(fn, ln),
                "course": course_code or "",
                "section": section or "",
                "time": tl_12,                   # show 12h
                "date": d.isoformat(),
                "start": st.strftime("%H:%M:%S"),
                "end": et.strftime("%H:%M:%S"),
                "faculty_id": fid,
            })
            time_set.add(tl_12)

    # --- REQUESTS: include course_code (visible time 12h) ---
    if include_requests:
        with connection.cursor() as cur:
            cur.execute("""
                SELECT 
                    u.utilization_id, u.date, u.start_time, u.end_time,
                    UPPER(LTRIM(RTRIM(u.status))) AS status,
                    fr.first_name, fr.last_name,
                    u.student_year_and_section, 
                    u.requested_by,
                    COALESCE(c1.course_code, c2.course_code) AS course_code
                FROM UTILIZATION_SLIP u
           LEFT JOIN FACULTY fr ON u.requested_by = fr.faculty_id
           LEFT JOIN LAB_SCHEDULE s ON u.schedule_id = s.schedule_id
           LEFT JOIN ASSIGNED_TEACHER at1 ON s.assigned_teacher_id = at1.assigned_teacher_id
           LEFT JOIN COURSE          c1   ON at1.course_id = c1.course_id
           LEFT JOIN ASSIGNED_TEACHER at2 ON u.assigned_teacher_id = at2.assigned_teacher_id
           LEFT JOIN COURSE          c2   ON at2.course_id = c2.course_id
               WHERE u.lab_id = %s
                 AND u.date BETWEEN %s AND %s
                 AND UPPER(LTRIM(RTRIM(u.status))) IN ('PENDING','FOR APPROVAL')
                 AND (u.schedule_id IS NULL OR UPPER(COALESCE(s.status,'')) IN ('REJECTED','CANCELLED',''))
               ORDER BY u.date, u.start_time
            """, [lab_id, start_w, end_w])
            seen = set()
            for uid, d, st, et, status, rfn, rln, section, requested_by, course_code in cur.fetchall():
                key = (d, st, et, requested_by, section or "")
                if key in seen:
                    continue
                seen.add(key)

                # fallback for course_code
                if not course_code and d:
                    fc_code, _fc_name = _resolve_course_for_slip_fallback(requested_by, d)
                    course_code = fc_code

                dayname = d.strftime("%A")
                tl_12 = f"{_fmt_hhmm_12(st)} - {_fmt_hhmm_12(et)}" if (st and et) else "—"
                if tl_12 != "—":
                    time_set.add(tl_12)
                req_entries[dayname][tl_12].append({
                    "id": uid,
                    "status": status or "PENDING",
                    "section": section or "",
                    "course": course_code or "",     # include code
                    "time": "" if tl_12 == "—" else tl_12,   # 12h visible
                    "date": d.isoformat(),
                    "start": st.strftime("%H:%M:%S") if st else "",
                    "end": et.strftime("%H:%M:%S") if et else "",
                    "requested_by": _short_name(rfn, rln),
                })

    return JsonResponse({
        "week_range": week_range_display,
        "entries": entries,
        "req_entries": req_entries,
        "time_slots": sorted(time_set),
    })

# ---------------------------------------------------------
# Optional: list all pending requests for a lab (JSON)
# ---------------------------------------------------------
@require_http_methods(["GET"])
def admin_pending_requests(request):
    lab_id = request.GET.get("lab_id")
    if not lab_id:
        return JsonResponse({"items": []})

    items = []
    with connection.cursor() as cur:
        cur.execute("""
            SELECT 
                u.utilization_id, u.date, u.start_time, u.end_time,
                u.student_year_and_section, u.status,
                f.first_name, f.last_name, l.lab_num,
                COALESCE(c1.course_code, c2.course_code) AS course_code
              FROM UTILIZATION_SLIP u
              JOIN FACULTY f ON f.faculty_id = u.requested_by AND ISNULL(f.is_archived,0)=0
         LEFT JOIN LABORATORIES l ON l.lab_id = u.lab_id
         LEFT JOIN LAB_SCHEDULE s ON u.schedule_id = s.schedule_id
         LEFT JOIN ASSIGNED_TEACHER at1 ON s.assigned_teacher_id = at1.assigned_teacher_id
         LEFT JOIN COURSE          c1   ON at1.course_id = c1.course_id
         LEFT JOIN ASSIGNED_TEACHER at2 ON u.assigned_teacher_id = at2.assigned_teacher_id
         LEFT JOIN COURSE          c2   ON at2.course_id = c2.course_id
             WHERE u.lab_id = %s
               AND UPPER(LTRIM(RTRIM(u.status))) IN ('PENDING','FOR APPROVAL')
             ORDER BY u.date DESC, u.start_time DESC
        """, [lab_id])
        for uid, d, st, et, sec, status, fn, ln, lab_num, course_code in cur.fetchall():
            items.append({
                "id": uid,
                "requestor": _short_name(fn, ln),
                "lab": f"Lab {lab_num}" if lab_num else "—",
                "day": d.strftime("%A") if d else "—",
                "date": d.isoformat() if d else None,
                "start_time": st.strftime("%H:%M:%S") if st else None,  # 24h for logic
                "end_time": et.strftime("%H:%M:%S") if et else None,
                "section": sec or "—",
                "course": (course_code or "—"),
                "status": status or "Pending",
            })

    return JsonResponse({"items": items})

# ---------------------------------------------------------
# Create repeating schedules (admin) — WITH operating_time_id
# ---------------------------------------------------------
@require_http_methods(["POST"])
def admin_create_schedule(request):
    if not request.session.get("user_id") or request.session.get("role") != "admin":
        return redirect("login")

    faculty_id = request.POST.get("faculty_id")
    course_id = request.POST.get("course_id")
    lab_id = request.POST.get("lab_id")
    section = request.POST.get("student_year_and_section")

    start_time_str = request.POST.get("start_time")
    end_time_str   = request.POST.get("end_time")
    date_start     = request.POST.get("start_date")
    date_end       = request.POST.get("end_date")
    selected_days  = request.POST.getlist("days")

    if not (faculty_id and course_id and lab_id and section and start_time_str and end_time_str and date_start and date_end and selected_days):
        return redirect(f"{reverse('manage_schedule')}?error=Missing+fields")

    try:
        start_date = dt.strptime(date_start, "%Y-%m-%d").date()
        end_date   = dt.strptime(date_end, "%Y-%m-%d").date()
    except Exception:
        return redirect(f"{reverse('manage_schedule')}?error=Invalid+dates")

    if start_date > end_date:
        return redirect(f"{reverse('manage_schedule')}?error=Start+date+must+be+before+end+date")

    try:
        t_start = dt.strptime(start_time_str, "%H:%M").time()
        t_end   = dt.strptime(end_time_str, "%H:%M").time()
    except Exception:
        return redirect(f"{reverse('manage_schedule')}?error=Invalid+time+format")

    if t_start >= t_end:
        return redirect(f"{reverse('manage_schedule')}?error=Start+time+must+be+before+end+time")

    day_map = {'Monday':0,'Tuesday':1,'Wednesday':2,'Thursday':3,'Friday':4,'Saturday':5,'Sunday':6}
    target_idx = {day_map[d] for d in selected_days if d in day_map}
    if not target_idx:
        return redirect(f"{reverse('manage_schedule')}?error=Select+at+least+one+day")

    created = 0
    skipped_closed, skipped_hours, skipped_assign, skipped_lab_conf, skipped_teacher_conf = [], [], [], [], []

    with connection.cursor() as cur:
        cur_date = start_date
        while cur_date <= end_date:
            if cur_date.weekday() not in target_idx:
                cur_date += timedelta(days=1)
                continue

            # Operating time (and ID)
            db_dow = _db_day_of_week(cur_date.weekday())
            cur.execute("""
                SELECT operating_time_id, is_open, start_time, end_time
                  FROM OPERATING_TIME
                 WHERE day_of_week = %s
            """, [db_dow])
            row = cur.fetchone()
            if not row:
                skipped_closed.append(cur_date); cur_date += timedelta(days=1); continue
            ot_id, is_open, ot_start, ot_end = row
            if not is_open:
                skipped_closed.append(cur_date); cur_date += timedelta(days=1); continue
            if not (ot_start <= t_start and t_end <= ot_end):
                skipped_hours.append(cur_date); cur_date += timedelta(days=1); continue

            # assigned_teacher_id
            cur.execute("""
                SELECT TOP 1 at.assigned_teacher_id
                  FROM ASSIGNED_TEACHER at
                  JOIN SEMESTER s ON s.semester_id = at.semester_id
                 WHERE at.faculty_id = %s
                   AND at.course_id  = %s
                   AND %s BETWEEN s.start_date AND s.end_date
                 ORDER BY s.start_date DESC
            """, [faculty_id, course_id, cur_date])
            at_row = cur.fetchone()
            if not at_row:
                skipped_assign.append(cur_date); cur_date += timedelta(days=1); continue
            assigned_teacher_id = at_row[0]

            # Lab conflict
            cur.execute("""
                SELECT COUNT(*) FROM LAB_SCHEDULE
                 WHERE lab_id=%s AND date=%s
                   AND NOT (end_time <= %s OR start_time >= %s)
            """, [lab_id, cur_date, start_time_str, end_time_str])
            if (cur.fetchone() or (0,))[0] > 0:
                skipped_lab_conf.append(cur_date); cur_date += timedelta(days=1); continue

            # Teacher conflict
            cur.execute("""
                SELECT COUNT(*)
                  FROM LAB_SCHEDULE s
                  JOIN ASSIGNED_TEACHER at ON s.assigned_teacher_id = at.assigned_teacher_id
                 WHERE at.faculty_id = %s
                   AND s.date = %s
                   AND NOT (s.end_time <= %s OR s.start_time >= %s)
            """, [faculty_id, cur_date, start_time_str, end_time_str])
            if (cur.fetchone() or (0,))[0] > 0:
                skipped_teacher_conf.append(cur_date); cur_date += timedelta(days=1); continue

            # Insert with operating_time_id
            cur.execute("""
                INSERT INTO LAB_SCHEDULE (
                    lab_id, assigned_teacher_id, reserved_to, date,
                    start_time, end_time, student_year_and_section, status, operating_time_id
                ) VALUES (%s,%s,%s,%s,%s,%s,%s,'Scheduled', %s)
            """, [lab_id, assigned_teacher_id, faculty_id, cur_date, start_time_str, end_time_str, section, ot_id])
            created += 1
            cur_date += timedelta(days=1)

    if created == 0:
        def _fmt(d: date) -> str: return d.strftime("%Y-%m-%d")
        parts = []
        if skipped_closed:       parts.append("Closed: " + ", ".join(_fmt(d) for d in skipped_closed))
        if skipped_hours:        parts.append("Out-of-hours: " + ", ".join(_fmt(d) for d in skipped_hours))
        if skipped_assign:       parts.append("No assignment: " + ", ".join(_fmt(d) for d in skipped_assign))
        if skipped_lab_conf:     parts.append("Lab conflicts: " + ", ".join(_fmt(d) for d in skipped_lab_conf))
        if skipped_teacher_conf: parts.append("Teacher conflicts: " + ", ".join(_fmt(d) for d in skipped_teacher_conf))
        msg = " | ".join(parts) if parts else "Nothing created."
        return redirect(f"{reverse('manage_schedule')}?error={msg}")

    skipped = len(skipped_closed)+len(skipped_hours)+len(skipped_assign)+len(skipped_lab_conf)+len(skipped_teacher_conf)
    msg = f"{created} slot(s) created"
    if skipped:
        msg += f"; {skipped} skipped"
    return redirect(f"{reverse('manage_schedule')}?success={msg}")

# ---------------------------------------------------------
# Approve / Reject (dedup + operating_time_id)
# ---------------------------------------------------------
@csrf_exempt
@require_http_methods(["POST", "GET"])
def approve_reservation(request, slip_id):
    # Try both keys, in case you sometimes use 'user_id'
    admin_id = request.session.get('user_id') or request.session.get('admin_id')
    role = (request.session.get('role') or '').lower()

    if not admin_id or role != 'admin':
        return redirect('login')

    with connection.cursor() as cur:
        cur.execute("""
            SELECT utilization_id, lab_id, date, start_time, end_time,
                   requested_by, student_year_and_section, status, schedule_id, assigned_teacher_id
              FROM UTILIZATION_SLIP
             WHERE utilization_id=%s
        """, [slip_id])
        r = cur.fetchone()
        if not r:
            return HttpResponseBadRequest("Reservation not found.")

        uid, lab_id, d, st, et, faculty_id, section, cur_status, existing_schedule_id, slip_at_id = r

        # requester active?
        cur.execute("SELECT ISNULL(is_archived,0), email FROM FACULTY WHERE faculty_id=%s", [faculty_id])
        rowf = cur.fetchone()
        if not rowf:
            return HttpResponseBadRequest("Requester not found.")
        if rowf[0] == 1:
            return HttpResponseBadRequest("Requester is archived.")
        requester_email = rowf[1]

        if cur_status in ('Rejected', 'Completed'):
            return HttpResponseBadRequest(f"Cannot approve: slip is {cur_status}.")

        # assigned teacher id (semester-aware)
        assigned_teacher_id = slip_at_id
        if not assigned_teacher_id:
            cur.execute("""
                SELECT TOP 1 at.assigned_teacher_id
                  FROM ASSIGNED_TEACHER at
                  JOIN SEMESTER s ON s.semester_id = at.semester_id
                 WHERE at.faculty_id = %s
                   AND %s BETWEEN s.start_date AND s.end_date
                 ORDER BY s.start_date DESC
            """, [faculty_id, d])
            at_row = cur.fetchone()
            assigned_teacher_id = at_row[0] if at_row else None

        # already linked + approved?
        if existing_schedule_id and cur_status == 'Approved':
            cur.execute("""UPDATE UTILIZATION_SLIP SET processed_by=%s WHERE utilization_id=%s""",
                        [admin_id, slip_id])
            return redirect('manage_schedule')

        # Try to reuse an identical schedule
        cur.execute("""
            SELECT TOP 1 schedule_id
              FROM LAB_SCHEDULE
             WHERE lab_id=%s AND date=%s AND start_time=%s AND end_time=%s
        """, [lab_id, d, st, et])
        sch_row = cur.fetchone()
        if sch_row:
            schedule_id = sch_row[0]
        else:
            # operating time id
            db_dow = _db_day_of_week(d.weekday())
            cur.execute("""
                SELECT operating_time_id, is_open, start_time, end_time
                  FROM OPERATING_TIME
                 WHERE day_of_week = %s
            """, [db_dow])
            ot_row = cur.fetchone()
            ot_id, is_open, ot_start, ot_end = (ot_row or (None, 0, None, None))
            if not ot_row or not is_open or not (ot_start <= st and et <= ot_end):
                return HttpResponseBadRequest("Outside operating hours.")

            # Lab conflict
            cur.execute("""
                SELECT COUNT(*)
                  FROM LAB_SCHEDULE
                 WHERE lab_id=%s AND date=%s
                   AND NOT (end_time <= %s OR start_time >= %s)
            """, [lab_id, d, st, et])
            if (cur.fetchone() or (0,))[0] > 0:
                return HttpResponseBadRequest("Conflict: another schedule overlaps this time in this lab.")

            # Create schedule (with operating_time_id)
            cur.execute("""
                INSERT INTO LAB_SCHEDULE (
                    lab_id, assigned_teacher_id, reserved_to, date, start_time, end_time,
                    student_year_and_section, status, operating_time_id
                ) OUTPUT INSERTED.schedule_id
                VALUES (%s,%s,%s,%s,%s,%s,%s,'Scheduled', %s)
            """, [lab_id, assigned_teacher_id, faculty_id, d, st, et, section, ot_id])
            schedule_id = cur.fetchone()[0]

        # Link & approve slip
        cur.execute("""
            UPDATE UTILIZATION_SLIP
               SET status='Approved',
                   processed_by=%s,
                   schedule_id=%s,
                   assigned_teacher_id = COALESCE(assigned_teacher_id, %s)
             WHERE utilization_id=%s
        """, [admin_id, schedule_id, assigned_teacher_id, slip_id])

        # Notify
        cur.execute("SELECT lab_num FROM LABORATORIES WHERE lab_id=%s", [lab_id])
        lab_num_row = cur.fetchone()
        lab_label = f"Lab {lab_num_row[0]}" if lab_num_row else f"Lab #{lab_id}"
        msg = f"Your lab request for {lab_label} on {d.strftime('%Y-%m-%d')} from {_fmt_hhmm(st)} to {_fmt_hhmm(et)} has been approved."
        cur.execute("""
            INSERT INTO NOTIFICATIONS (message, status, receiver_teacher_id, sender_teacher_id, created_at)
            VALUES (%s, 'Unread', %s, %s, GETDATE())
        """, [msg, faculty_id, admin_id])

        # 🔔 Realtime: push updated unread count to the requester
        cur.execute("""
            SELECT COUNT(*)
            FROM NOTIFICATIONS
            WHERE receiver_teacher_id = %s AND status = 'Unread'
        """, [faculty_id])
        cnt_row = cur.fetchone()
        unread = cnt_row[0] if cnt_row else 0

        cur.execute("""
            SELECT role
            FROM FACULTY
            WHERE faculty_id = %s AND ISNULL(is_archived,0) = 0
        """, [faculty_id])
        role_row = cur.fetchone()
        if role_row:
            role = (role_row[0] or "").strip().lower()
            channel = None
            if role == "teacher":
                channel = f"teacher-{faculty_id}"
            elif role == "admin":
                channel = f"admin-{faculty_id}"

            if channel:
                try:
                    realtime.trigger(channel, "notification", {
                        "unread_count": unread,
                    })
                except Exception as e:
                    print("Pusher error (admin approve_reservation):", e)

        if requester_email:
            try:
                send_mail("Lab Request Approved", msg, settings.DEFAULT_FROM_EMAIL, [requester_email], fail_silently=True)
            except Exception:
                pass

    return redirect('manage_schedule')


@csrf_exempt
@require_http_methods(["POST", "GET"])
def reject_reservation(request, slip_id):
    admin_id = request.session.get('user_id') or request.session.get('admin_id')
    role = (request.session.get('role') or '').lower()

    if not admin_id or role != 'admin':
        return redirect('login')

    with connection.cursor() as cur:
        cur.execute("""
            SELECT lab_id, date, start_time, end_time, requested_by, status
              FROM UTILIZATION_SLIP
             WHERE utilization_id=%s
        """, [slip_id])
        r = cur.fetchone()
        if not r:
            return HttpResponseBadRequest("Reservation not found.")
        lab_id, d, st, et, faculty_id, cur_status = r

        if cur_status in ('Rejected','Completed'):
            return HttpResponseBadRequest(f"Cannot reject: slip is {cur_status}.")

        cur.execute("SELECT ISNULL(is_archived,0), email FROM FACULTY WHERE faculty_id=%s", [faculty_id])
        frow = cur.fetchone()
        if not frow or frow[0] == 1:
            return HttpResponseBadRequest("Requester not found or archived.")

        cur.execute("""
            UPDATE UTILIZATION_SLIP
               SET status='Rejected', processed_by=%s
             WHERE utilization_id=%s
        """, [admin_id, slip_id])

        cur.execute("SELECT lab_num FROM LABORATORIES WHERE lab_id=%s", [lab_id])
        lab_num_row = cur.fetchone()
        lab_label = f"Lab {lab_num_row[0]}" if lab_num_row else f"Lab #{lab_id}"
        msg = f"Your lab request for {lab_label} on {d.strftime('%Y-%m-%d')} from {_fmt_hhmm(st)} to {_fmt_hhmm(et)} has been rejected."

        cur.execute("""
            INSERT INTO NOTIFICATIONS (message, status, receiver_teacher_id, sender_teacher_id, created_at)
            VALUES (%s, 'Unread', %s, %s, GETDATE())
        """, [msg, faculty_id, admin_id])

        # 🔔 Realtime: push updated unread count to the requester
        cur.execute("""
            SELECT COUNT(*)
            FROM NOTIFICATIONS
            WHERE receiver_teacher_id = %s AND status = 'Unread'
        """, [faculty_id])
        cnt_row = cur.fetchone()
        unread = cnt_row[0] if cnt_row else 0

        cur.execute("""
            SELECT role
            FROM FACULTY
            WHERE faculty_id = %s AND ISNULL(is_archived,0) = 0
        """, [faculty_id])
        role_row = cur.fetchone()
        if role_row:
            role = (role_row[0] or "").strip().lower()
            channel = None
            if role == "teacher":
                channel = f"teacher-{faculty_id}"
            elif role == "admin":
                channel = f"admin-{faculty_id}"

            if channel:
                try:
                    realtime.trigger(channel, "notification", {
                        "unread_count": unread,
                    })
                except Exception as e:
                    print("Pusher error (admin reject_reservation):", e)

        requester_email = frow[1]
        if requester_email:
            try:
                send_mail("Lab Request Rejected", msg, settings.DEFAULT_FROM_EMAIL, [requester_email], fail_silently=True)
            except Exception:
                pass

    return redirect('manage_schedule')

def login_view(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')

        # FACULTY login only (STUDENTS login removed)
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT faculty_id, LOWER(ISNULL(role,'teacher')) AS role, password
                FROM FACULTY
                WHERE email=%s AND ISNULL(is_archived,0)=0
            """, [email])
            faculty = cursor.fetchone()

        if faculty and check_password(password, faculty[2]):
            faculty_id, role, _ = faculty
            request.session['user_id'] = faculty_id
            request.session['role'] = role

            # superadmin uses the same dashboard as admin
            if role == 'admin':
                return redirect('dashboard')
            elif role == 'teacher':
                return redirect('teacher_dashboard')

        # Failed login
        return render(request, 'admin/login.html', {'error': 'Invalid email or password'})

    return render(request, 'admin/login.html')


@csrf_exempt
def mark_notifications_read(request):
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request"}, status=405)
    admin_id = request.session.get("user_id")
    if not admin_id or request.session.get("role") != "admin":
        return JsonResponse({"error": "Unauthorized"}, status=401)

    with connection.cursor() as cursor:
        # NOTE: If your table has a separate receiver column for admins, change this WHERE clause.
        cursor.execute("""
            UPDATE NOTIFICATIONS
            SET status = 'Read'
            WHERE receiver_teacher_id = %s AND status = 'Unread'
        """, [admin_id])
    return JsonResponse({"status": "success"})

@require_GET
def admin_notification_detail(request, notification_id: int):
    admin_id = request.session.get("user_id")
    if not admin_id or request.session.get("role") != "admin":
        return JsonResponse({"error": "Unauthorized"}, status=401)

    with connection.cursor() as c:
        c.execute("""
            SELECT message
            FROM NOTIFICATIONS
            WHERE notification_id = %s AND receiver_teacher_id = %s
        """, [notification_id, admin_id])
        row = c.fetchone()

    if not row:
        return JsonResponse({"error": "Not found"}, status=404)
    return JsonResponse({"full_message": row[0]})

# OPTIONAL: only if you later add a Load More button
@require_GET
def admin_notifications_load_more(request):
    admin_id = request.session.get("user_id")
    if not admin_id or request.session.get("role") != "admin":
        return JsonResponse({"error": "Unauthorized"}, status=401)

    last_id = request.GET.get("last_id")
    page_size = 10

    with connection.cursor() as c:
        # compute OFFSET based on created_at position of last_id
        offset = 0
        if last_id:
            c.execute("""
                SELECT COUNT(*) FROM NOTIFICATIONS
                WHERE receiver_teacher_id = %s
                  AND created_at > (SELECT created_at FROM NOTIFICATIONS WHERE notification_id = %s)
            """, [admin_id, last_id])
            offset = c.fetchone()[0]

        c.execute("""
            SELECT notification_id, message, status, created_at
            FROM NOTIFICATIONS
            WHERE receiver_teacher_id = %s
            ORDER BY created_at DESC
            OFFSET %s ROWS FETCH NEXT %s ROWS ONLY
        """, [admin_id, offset, page_size])
        rows = c.fetchall() or []

    data = [{
        "id": r[0],
        "message": r[1],
        "status": r[2],
        "created_at": r[3].strftime("%b %d, %Y %H:%M") if r[3] else ""
    } for r in rows]

    return JsonResponse({"notifications": data, "no_more": len(data) < page_size})


def logout_view(request):
    logout(request)
    request.session.flush()
    return redirect('login')


def test_db_connection(request):
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT 1")
            row = cursor.fetchone()
        return HttpResponse(f"✅ Database connected! Result: {row}")
    except Exception as e:
        return HttpResponse(f"❌ Database connection failed: {e}")


def test_email(request):
    try:
        send_mail(
            subject='Test Email from SmartLab',
            message='This is a test email to confirm your SMTP settings are working.',
            from_email=None,  # Uses DEFAULT_FROM_EMAIL from settings
            recipient_list=['Seanzed08@gmail.com'],  # Replace with your real email
            fail_silently=False,
        )
        return HttpResponse("✅ Email sent successfully!")
    except Exception as e:
        return HttpResponse(f"❌ Failed to send email: {e}")


@csrf_exempt
def rfid_scan(request):
    if request.method == 'POST':
        rfid_code = request.POST.get('rfid_code')
        lab_id = request.POST.get('lab_id')

        with connection.cursor() as cursor:
            cursor.execute("SELECT faculty_id FROM FACULTY WHERE rfid_sticker_id = %s AND is_archived = 0", [rfid_code])
            row = cursor.fetchone()
            if not row:
                return HttpResponse("Invalid RFID code", status=404)
            faculty_id = row[0]

            cursor.execute("""
                SELECT 1
                FROM LABORATORIES
                WHERE lab_id = %s AND faculty_id = %s
            """, [lab_id, faculty_id])
            if cursor.fetchone() is None:
                return HttpResponse("You are not assigned to this lab.", status=403)

            now = timezone.localtime().time()
            today = timezone.localdate()

            cursor.execute("""
                SELECT schedule_id, assigned_teacher_id
                FROM LAB_SCHEDULE
                WHERE lab_id = %s
                  AND reserved_to = %s
                  AND date = %s
                  AND start_time <= %s
                  AND end_time >= %s
            """, [lab_id, faculty_id, today, now, now])
            sched_row = cursor.fetchone()

            if not sched_row:
                return HttpResponse("No scheduled class found for this time.")

            schedule_id, assigned_teacher_id = sched_row

            cursor.execute("SELECT COUNT(*) FROM UTILIZATION_SLIP WHERE schedule_id = %s AND date = %s", [schedule_id, today])
            exists = cursor.fetchone()[0]

            semester_id = None
            if assigned_teacher_id:
                cursor.execute("SELECT semester_id FROM ASSIGNED_TEACHER WHERE assigned_teacher_id = %s", [assigned_teacher_id])
                at_row = cursor.fetchone()
                if at_row:
                    semester_id = at_row[0]

            if semester_id is None:
                cursor.execute("""
                    SELECT TOP 1 semester_id
                    FROM SEMESTER
                    WHERE start_date <= %s AND end_date >= %s
                    ORDER BY start_date DESC
                """, [today, today])
                sem = cursor.fetchone()
                semester_id = sem[0] if sem else None

            if exists == 0:
                cursor.execute("""
                    INSERT INTO UTILIZATION_SLIP (
                        schedule_id, date, time_duration, start_time, end_time,
                        requested_by, student_year_and_section, number_of_students,
                        remarks, processed_by, status, created_at, lab_id, semester_id
                    ) VALUES (
                        %s, %s, %s, %s, %s,
                        %s, %s, %s,
                        %s, %s, %s, GETDATE(), %s, %s
                    )
                """, [
                    schedule_id, today, time(0, 0), None, None,
                    faculty_id, '', 0,
                    'RFID quick-log', None, 'Approved', lab_id, semester_id
                ])

            return HttpResponse("Utilization slip logged successfully.")

    return HttpResponse("Only POST allowed.", status=405)


def rfid_scan_page(request):
    return render(request, "admin/scan_rfid.html", {
        "page_title": "RFID Scan Page"
    })
# --- debug helper for RFID schedule matching ---
def _debug_no_match(cursor, *, lab_id, teacher_id, today, current_time, grace_minutes):
    from datetime import datetime as dt, timedelta

    # nearest schedule today (diagnostic)
    cursor.execute("""
        SELECT TOP 1 s.schedule_id, s.start_time, s.end_time, s.assigned_teacher_id, s.reserved_to
        FROM LAB_SCHEDULE s
        WHERE s.lab_id=%s AND s.date=%s
        ORDER BY
          CASE WHEN s.start_time <= %s AND %s <= DATEADD(MINUTE, %s, s.end_time) THEN 0 ELSE 1 END,
          ABS(DATEDIFF(SECOND, s.start_time, %s))
    """, [lab_id, today, current_time, current_time, grace_minutes, current_time])
    near = cursor.fetchone()
    nearest = None
    outside_by_min = None
    if near:
        sid, st, et, atid, rto = near
        cursor.execute("SELECT faculty_id FROM ASSIGNED_TEACHER WHERE assigned_teacher_id=%s", [atid])
        at_fac = cursor.fetchone()
        nearest = {
            "schedule_id": sid,
            "start": str(st),
            "end": str(et),
            "assigned_teacher_id": atid,
            "assigned_faculty_id": (at_fac[0] if at_fac else None),
            "reserved_to": rto,
        }
        cursor.execute("SELECT DATEDIFF(MINUTE, DATEADD(MINUTE, %s, %s), %s)",
                       [grace_minutes, et, current_time])
        (outside_by_min,) = cursor.fetchone() or (None,)

    # count matches by assigned teacher path
    cursor.execute("""
        SELECT COUNT(*)
        FROM LAB_SCHEDULE s
        JOIN ASSIGNED_TEACHER at ON at.assigned_teacher_id = s.assigned_teacher_id
        WHERE s.lab_id=%s AND s.date=%s AND at.faculty_id=%s
          AND DATEADD(MINUTE,-10,s.start_time) <= %s
          AND %s <= DATEADD(MINUTE,%s,s.end_time)
    """, [lab_id, today, teacher_id, current_time, current_time, grace_minutes])
    (by_assigned,) = cursor.fetchone() or (0,)

    # count matches by legacy reserved_to path
    cursor.execute("""
        SELECT COUNT(*)
        FROM LAB_SCHEDULE s
        WHERE s.lab_id=%s AND s.date=%s AND s.reserved_to=%s
          AND DATEADD(MINUTE,-10,s.start_time) <= %s
          AND %s <= DATEADD(MINUTE,%s,s.end_time)
    """, [lab_id, today, teacher_id, current_time, current_time, grace_minutes])
    (by_reserved,) = cursor.fetchone() or (0,)

    return {
        "lab_id": lab_id,
        "teacher_id": teacher_id,
        "today": str(today),
        "current_time": str(current_time),
        "grace_minutes": grace_minutes,
        "matches_by_assigned_teacher": by_assigned,
        "matches_by_reserved_to": by_reserved,
        "nearest_schedule_today": nearest,
        "outside_minutes_after_end_plus_grace": outside_by_min,
    }

from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.db import connection
from django.utils import timezone

def _column_exists(cursor, table_name: str, column_name: str) -> bool:
    cursor.execute("""
        SELECT 1
        FROM INFORMATION_SCHEMA.COLUMNS
        WHERE TABLE_NAME = %s AND COLUMN_NAME = %s
    """, [table_name, column_name])
    return cursor.fetchone() is not None

def _bool_as_int(b):
    return 1 if b else 0

def _complete_overdue_sessions(cursor, today, now_time):
    """
    For UTILIZATION_SLIP rows (today) still 'Active' but with end_time <= now_time:
      1) auto time_out students to the slip's end_time
      2) mark the slip as 'Completed'
    """
    # Auto student tap-out to session end
    cursor.execute("""
        UPDATE cla
           SET time_out = u.end_time
          FROM COMPUTER_LAB_ATTENDANCE AS cla
          JOIN UTILIZATION_SLIP       AS u
            ON u.utilization_id = cla.utilization_id
         WHERE u.status   = 'Active'
           AND u.date     = %s
           AND u.end_time <= %s
           AND cla.time_out IS NULL
    """, [today, now_time])

    # Mark sessions completed
    cursor.execute("""
        UPDATE UTILIZATION_SLIP
           SET status = 'Completed'
         WHERE status   = 'Active'
           AND date     = %s
           AND end_time <= %s
    """, [today, now_time])

# Build a safe INSERT for UTILIZATION_SLIP that only includes columns that exist.
def _insert_utilization_for_schedule(cursor, schedule_id, today, actual_start, scheduled_end,
                                     teacher_id, time_duration):
    """
    Creates an Active UTILIZATION_SLIP row for the given LAB_SCHEDULE.
    Column list is computed at runtime to avoid 42S22 if your DB is missing some columns.
    Remarks will be left NULL (no 'Auto RFID Check-in' text).
    """
    has_us_num_students  = _column_exists(cursor, "UTILIZATION_SLIP", "number_of_students")
    has_us_assigned_tid  = _column_exists(cursor, "UTILIZATION_SLIP", "assigned_teacher_id")
    has_ls_num_students  = _column_exists(cursor, "LAB_SCHEDULE",    "number_of_students")

    # Base column list
    cols = [
        "schedule_id", "date", "time_duration", "start_time", "end_time",
        "requested_by", "student_year_and_section", "remarks", "processed_by",
        "status", "created_at", "lab_id"
    ]
    if has_us_num_students:
        # insert before remarks
        cols.insert(7, "number_of_students")
    if has_us_assigned_tid:
        cols.append("assigned_teacher_id")

    # Build SELECT part
    select_parts = [
        "%s",             # schedule_id
        "%s",             # date
        "%s",             # time_duration
        "%s",             # start_time
        "s.end_time",     # end_time (from schedule)
        "%s",             # requested_by (teacher_id)
        "s.student_year_and_section",
        # (optionally number_of_students here)
        "%s",             # remarks (now NULL via param)
        "at.assigned_by", # processed_by
        "'Active'",       # status (constant)
        "GETDATE()",      # created_at
        "s.lab_id"        # lab_id
        # (optionally assigned_teacher_id at the end)
    ]

    # If UTILIZATION_SLIP has number_of_students, pull from LAB_SCHEDULE if present
    if has_us_num_students:
        num_students_expr = "s.number_of_students" if has_ls_num_students else "NULL"
        # Insert in the same position as in cols → before remarks
        # In select_parts, that's index 7 (right before remarks %s)
        select_parts.insert(7, num_students_expr)

    if has_us_assigned_tid:
        select_parts.append("s.assigned_teacher_id")

    col_list = ", ".join(cols)
    select_clause = ", ".join(select_parts)

    sql = f"""
        INSERT INTO UTILIZATION_SLIP ({col_list})
        SELECT {select_clause}
          FROM LAB_SCHEDULE s
          LEFT JOIN ASSIGNED_TEACHER at ON at.assigned_teacher_id = s.assigned_teacher_id
         WHERE s.schedule_id = %s
    """

    # ORDER OF PARAMS MUST MATCH %s IN select_parts
    params = [
        schedule_id,   # schedule_id
        today,         # date
        time_duration, # time_duration
        actual_start,  # start_time
        # s.end_time (no param)
        teacher_id,    # requested_by
        # s.student_year_and_section (no param)
        # number_of_students (no param – it's a column or NULL)
        None,          # remarks → will be stored as NULL
    ]

    # WHERE s.schedule_id = %s
    params.append(schedule_id)

    cursor.execute(sql, params)


from teacher.realtime import realtime  # ✅ NEW
from datetime import time as dtime 
@csrf_exempt
def rfid_check_in(request):
    if request.method != "POST":
        return JsonResponse({"error": "Invalid method."}, status=405)

    rfid_code = request.POST.get("rfid_code")
    reader_id = request.POST.get("rfid_reader_id")

    now = timezone.localtime()
    today = now.date()
    current_time = now.time()

    with connection.cursor() as cursor:
        # ✅ Always clean up overdue Active sessions first
        _complete_overdue_sessions(cursor, today, current_time)

        # Teacher?
        cursor.execute("""
            SELECT faculty_id, first_name, last_name
              FROM FACULTY
             WHERE rfid_sticker_id = %s AND is_archived = 0
        """, [rfid_code])
        teacher = cursor.fetchone()
        if teacher:
            return handle_teacher_rfid(cursor, teacher, reader_id, now, today, current_time)

        # Student?
        cursor.execute("""
            SELECT student_id, first_name, last_name
              FROM STUDENTS
             WHERE rfid_sticker_id = %s AND is_archived = 0
        """, [rfid_code])
        student = cursor.fetchone()
        if student:
            return handle_student_rfid(cursor, student, reader_id, now, today, current_time)

        return JsonResponse({"error": "Unknown RFID code."}, status=404)


# -------------------------------------------------
# Teacher flow
# -------------------------------------------------
def handle_teacher_rfid(cursor, teacher, reader_id, now, today, current_time):
    from datetime import datetime as dt

    teacher_id, fname, lname = teacher

    # Reader → lab
    cursor.execute("SELECT lab_id, lab_num FROM LABORATORIES WHERE rfid_reader_id = %s", [reader_id])
    lab_row = cursor.fetchone()
    if not lab_row:
        return JsonResponse({"error": "Unregistered RFID reader."}, status=404)
    lab_id, lab_num = lab_row

    # Match a valid schedule for this teacher
    cursor.execute("""
        SELECT TOP 1
            s.schedule_id, s.start_time, s.end_time, s.assigned_teacher_id,
            s.reserved_to, at.faculty_id AS at_faculty_id
        FROM LAB_SCHEDULE AS s
        LEFT JOIN ASSIGNED_TEACHER AS at
               ON at.assigned_teacher_id = s.assigned_teacher_id
        WHERE s.lab_id = %s
          AND s.date   = %s
          AND DATEADD(MINUTE, -10, s.start_time) <= %s
          AND %s <= s.end_time
          AND (s.reserved_to = %s OR at.faculty_id = %s)
        ORDER BY s.start_time
    """, [lab_id, today, current_time, current_time, teacher_id, teacher_id])

    schedule = cursor.fetchone()

    if not schedule:
        return JsonResponse({
            "error": "No active schedule or approved request at this time.",
            "role": "teacher",
            "teacher": f"{fname} {lname}",
            "lab": f"Lab {lab_num}",
        }, status=403)

    schedule_id, scheduled_start, scheduled_end, assigned_teacher_id, _reserved_to, _at_fac = schedule

    # Get processed_by
    cursor.execute("SELECT assigned_by FROM ASSIGNED_TEACHER WHERE assigned_teacher_id = %s", [assigned_teacher_id])
    row = cursor.fetchone()
    processed_by = row[0] if row else None

    # Check if a slip already exists
    cursor.execute("""
        SELECT TOP 1 utilization_id, start_time, end_time, status
        FROM UTILIZATION_SLIP
        WHERE schedule_id=%s AND date=%s AND requested_by=%s
        ORDER BY created_at DESC
    """, [schedule_id, today, teacher_id])

    existing = cursor.fetchone()

    # ----------------------------------------
    # CASE 1 — Already Completed
    # ----------------------------------------
    if existing:
        utilization_id, start_in, _end_orig, status = existing

        if status == "Completed":
            return JsonResponse({
                "status": "Already tapped out.",
                "role": "teacher",
                "teacher": f"{fname} {lname}",
                "lab": f"Lab {lab_num}",
            })

        # ----------------------------------------
        # CASE 2 — Approved → Active (Check-in)
        # ----------------------------------------
        if status == "Approved":
            duration = dt.combine(today, scheduled_end) - dt.combine(today, current_time)
            new_dur = (dt.min + duration).time()

            cursor.execute("""
                UPDATE UTILIZATION_SLIP
                   SET start_time=%s, end_time=%s, time_duration=%s, status='Active'
                 WHERE utilization_id=%s
            """, [current_time, scheduled_end, new_dur, utilization_id])

            # 🔔 Push updated active sessions (REALTIME)
            _push_active_sessions_for_lab(cursor, lab_id)

            return JsonResponse({
                "status": "Check-in successful (from approved request).",
                "role": "teacher",
                "teacher": f"{fname} {lname}",
                "lab": f"Lab {lab_num}",
                "checked_in_at": now.strftime("%H:%M:%S"),
            })

        # ----------------------------------------
        # CASE 3 — Active → Completed (Tap-out)
        # ----------------------------------------
        if status == "Active":
            duration = dt.combine(today, current_time) - dt.combine(today, start_in)
            new_dur = (dt.min + duration).time()

            cursor.execute("""
                UPDATE UTILIZATION_SLIP
                   SET end_time=%s, time_duration=%s, status='Completed'
                 WHERE utilization_id=%s
            """, [current_time, new_dur, utilization_id])

            cursor.execute("""
                UPDATE COMPUTER_LAB_ATTENDANCE
                   SET time_out=%s
                 WHERE utilization_id=%s AND time_out IS NULL
            """, [current_time, utilization_id])

            # 🔔 Push updated active sessions (REALTIME)
            _push_active_sessions_for_lab(cursor, lab_id)

            return JsonResponse({
                "status": "Tap-out successful.",
                "role": "teacher",
                "teacher": f"{fname} {lname}",
                "lab": f"Lab {lab_num}",
                "actual_end": now.strftime("%H:%M:%S"),
            })

    # ----------------------------------------
    # CASE 4 — No slip exists → New slip ACTIVE
    # ----------------------------------------
    actual_start = current_time
    duration = dt.combine(today, scheduled_end) - dt.combine(today, actual_start)
    time_dur = (dt.min + duration).time()

    _insert_utilization_for_schedule(
        cursor=cursor,
        schedule_id=schedule_id,
        today=today,
        actual_start=actual_start,
        scheduled_end=scheduled_end,
        teacher_id=teacher_id,
        time_duration=time_dur
    )

    # 🔔 Push updated active sessions (REALTIME)
    _push_active_sessions_for_lab(cursor, lab_id)

    return JsonResponse({
        "status": "Check-in successful.",
        "role": "teacher",
        "teacher": f"{fname} {lname}",
        "lab": f"Lab {lab_num}",
        "checked_in_at": now.strftime("%H:%M:%S"),
    })


# -------------------------------------------------
# Student flow
# -------------------------------------------------
def handle_student_rfid(cursor, student, reader_id, now, today, current_time):
    from datetime import datetime as dt, timedelta

    student_id, sfname, slname = student

    # Map reader -> lab
    cursor.execute("SELECT lab_id, lab_num FROM LABORATORIES WHERE rfid_reader_id = %s", [reader_id])
    lab_row = cursor.fetchone()
    if not lab_row:
        return JsonResponse({"error": "Unregistered RFID reader."}, status=404)
    lab_id, lab_num = lab_row

    LATE_MIN = 0  # set to 5 if you want exit grace

    # Find an ACTIVE overlapping session
    cursor.execute(f"""
        SELECT TOP 1 utilization_id, date, start_time, end_time, requested_by
          FROM UTILIZATION_SLIP
         WHERE lab_id = %s
           AND date   = %s
           AND status = 'Active'
           AND start_time <= %s
           AND %s <= DATEADD(MINUTE, {LATE_MIN}, end_time)
         ORDER BY start_time
    """, [lab_id, today, current_time, current_time])
    util = cursor.fetchone()

    if not util:
        # (keep your debug block if you had one)
        return JsonResponse({
            "error": "No active session in this lab.",
            "role": "student", "student": f"{sfname} {slname}", "lab": f"Lab {lab_num}",
        }, status=403)

    utilization_id, udate, ust, uet, req_by = util

    # Did this student already tap for this utilization?
    cursor.execute("""
        SELECT TOP 1 attendance_sheet_id, time_in, time_out
          FROM COMPUTER_LAB_ATTENDANCE
         WHERE student_id = %s AND utilization_id = %s
         ORDER BY attendance_sheet_id DESC
    """, [student_id, utilization_id])
    row = cursor.fetchone()

    if row:
        attendance_id, time_in, time_out = row
        if time_out is None:
            cursor.execute("""
                UPDATE COMPUTER_LAB_ATTENDANCE
                   SET time_out = %s
                 WHERE attendance_sheet_id = %s
            """, [current_time, attendance_id])
            return JsonResponse({
                "status": "Student tap-out recorded.",
                "role": "student", "student": f"{sfname} {slname}", "lab": f"Lab {lab_num}",
                "time_out": now.strftime("%H:%M:%S")
            })
        else:
            return JsonResponse({
                "status": "Already tapped out.",
                "role": "student", "student": f"{sfname} {slname}", "lab": f"Lab {lab_num}"
            })

    # First tap → insert
    cursor.execute("""
        INSERT INTO COMPUTER_LAB_ATTENDANCE (student_id, utilization_id, time_in, remarks, created_at)
        VALUES (%s, %s, %s, %s, %s)
    """, [student_id, utilization_id, current_time, 'RFID tap-in', today])

    return JsonResponse({
        "status": "Student tap-in recorded.",
        "role": "student", "student": f"{sfname} {slname}", "lab": f"Lab {lab_num}",
        "time_in": now.strftime("%H:%M:%S")
    })


# -------------------------------------------------
# Active sessions (polled by frontend)
# -------------------------------------------------
@csrf_exempt
def get_active_sessions(request):
    lab_id = request.GET.get("lab_id")
    if not lab_id:
        return JsonResponse({"active_sessions": []})

    now = timezone.localtime()
    today = now.date()
    now_t = now.time()

    with connection.cursor() as cursor:
        # ✅ Keep sessions tidy on each poll
        _complete_overdue_sessions(cursor, today, now_t)

        cursor.execute("""
            SELECT date, start_time, end_time, requested_by
              FROM UTILIZATION_SLIP
             WHERE lab_id = %s AND status = 'Active' AND date = %s
        """, [lab_id, today])
        rows = cursor.fetchall()

    return JsonResponse({
        "active_sessions": [
            {
                "date":       r[0].isoformat(),
                "start_time": (r[1].strftime("%H:%M:%S") if r[1] else "00:00:00"),
                "end_time":   (r[2].strftime("%H:%M:%S") if r[2] else "00:00:00"),
                "faculty_id": r[3],
            }
            for r in rows
        ]
    })



@csrf_exempt
def rfid_test_page(request):
    response_data = None

    if request.method == "POST":
        rfid_code = request.POST.get("rfid_code")
        rfid_reader_id = request.POST.get("rfid_reader_id")
        payload = {
            "rfid_code": rfid_code,
            "rfid_reader_id": rfid_reader_id
        }

        response = requests.post("http://127.0.0.1:8000/rfid/check-in/", data=payload)
        try:
            response_data = response.json()
        except:
            response_data = {"error": "Invalid response from endpoint."}

    return render(request, "admin/rfid_test_page.html", {"response_data": response_data})

def admin_profile(request):
    """
    Admin can update only their name, email, password, and profile image.
    Role and RFID are read-only (displayed only).
    Shows a one-time toast after successful update via `profile_toast`.
    """
    if not request.session.get('user_id') or request.session.get('role') != 'admin':
        return redirect('login')

    admin_id = request.session.get('user_id')
    # use your header builder; if you use get_admin_header_context, swap it here:
    context = build_user_header_context(admin_id)
    context["current_page"] = "My Profile"

    # Load current admin record
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT faculty_id, first_name, middle_name, last_name, email, password, role,
                   rfid_sticker_id, profile_image
              FROM FACULTY
             WHERE faculty_id = %s
        """, [admin_id])
        admin = cursor.fetchone()

    if not admin:
        return HttpResponse("Admin not found.")

    if request.method == "POST":
        first_name = (request.POST.get("first_name") or "").strip()
        middle_name = (request.POST.get("middle_name") or "").strip()
        last_name  = (request.POST.get("last_name") or "").strip()
        email      = (request.POST.get("email") or "").strip()
        password_input = (request.POST.get("password") or "").strip()

        # Keep existing hashed password if field left blank
        password_hashed = make_password(password_input) if password_input else admin[5]

        profile_image_path = admin[8]  # keep the old one by default

        # Optional upload
        uploaded_file = request.FILES.get('profile_image')
        if uploaded_file:
            # (simple guard – skip heavy validation)
            upload_dir = os.path.join(settings.MEDIA_ROOT, 'faculty_images')
            os.makedirs(upload_dir, exist_ok=True)
            filename = f"{admin_id}_{uploaded_file.name}"
            filepath = os.path.join(upload_dir, filename)
            with open(filepath, 'wb+') as dest:
                for chunk in uploaded_file.chunks():
                    dest.write(chunk)
            profile_image_path = f"faculty_images/{filename}"

        try:
            with connection.cursor() as cursor:
                cursor.execute("""
                    UPDATE FACULTY
                       SET first_name=%s,
                           middle_name=%s,
                           last_name=%s,
                           email=%s,
                           password=%s,
                           profile_image=%s
                     WHERE faculty_id=%s
                """, [
                    first_name, middle_name, last_name,
                    email, password_hashed,
                    profile_image_path, admin_id
                ])

            # ✅ set a one-time toast for the next GET render
            request.session["profile_toast"] = {
                "level": "success",
                "text": "Profile updated successfully."
            }
        except Exception as e:
            request.session["profile_toast"] = {
                "level": "error",
                "text": f"Update failed: {e}"
            }

        return redirect('admin_profile')  

    
    context.update({
        "admin_id": admin[0],
        "first_name": admin[1],
        "middle_name": admin[2],
        "last_name": admin[3],
        "email": admin[4],
        "password": "",        
        "role": admin[6],       
        "rfid": admin[7],         
        "profile_image": admin[8],
    })

 
    context["profile_toast"] = request.session.pop("profile_toast", None)

    return render(request, "admin/profile.html", context)

def _time_hhmm(t):
    if not t:
        return ""
    s = str(t)
    return s[:5]

@transaction.atomic
def manage_operating_time(request):
    """
    Admin page to manage weekly operating hours stored in OPERATING_TIME:
      day_of_week TINYINT (0=Sun..6=Sat)
      is_open     BIT
      start_time  TIME NULL
      end_time    TIME NULL
    """
    if not request.session.get('user_id') or request.session.get('role') != 'admin':
        return redirect('login')

    admin_id = request.session.get('user_id')
    context = build_user_header_context(admin_id)
    context["current_page"] = "Operating Time"

 
    try:
        with connection.cursor() as c:
            c.execute("SELECT 1 FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME = 'OPERATING_TIME'")
            if c.fetchone() is None:
                return HttpResponseBadRequest("OPERATING_TIME table is missing. Please create it first.")
    except Exception:
        pass

    days_order = [
        (1, "Monday"),
        (2, "Tuesday"),
        (3, "Wednesday"),
        (4, "Thursday"),
        (5, "Friday"),
        (6, "Saturday"),
        (0, "Sunday"),
    ]
    default_open, default_close = "07:00", "21:00"

    if request.method == "POST":
        with connection.cursor() as cursor:
            for dow, _ in days_order:
                is_open = 1 if request.POST.get(f"is_open_{dow}") in ("on", "1", "true", "True") else 0
                start_t = (request.POST.get(f"open_{dow}") or "").strip()
                end_t   = (request.POST.get(f"close_{dow}") or "").strip()

                if is_open:
                    if not start_t or not end_t:
                        messages.error(request, f"Please set both start and end time for {dow}.")
                        return redirect("manage_operating_time")
                    if start_t >= end_t:
                        messages.error(request, "Start time must be earlier than end time.")
                        return redirect("manage_operating_time")
                    start_val, end_val = start_t, end_t
                else:
                    start_val, end_val = None, None

                cursor.execute("SELECT operating_time_id FROM OPERATING_TIME WHERE day_of_week = %s", [dow])
                row = cursor.fetchone()
                if row:
                    cursor.execute("""
                        UPDATE OPERATING_TIME
                           SET is_open = %s,
                               start_time = %s,
                               end_time = %s
                         WHERE day_of_week = %s
                    """, [is_open, start_val, end_val, dow])
                else:
                    cursor.execute("""
                        INSERT INTO OPERATING_TIME (day_of_week, is_open, start_time, end_time)
                        VALUES (%s, %s, %s, %s)
                    """, [dow, is_open, start_val, end_val])

        messages.success(request, "Operating time updated.")
        return redirect("manage_operating_time")

   
    existing = {}
    with connection.cursor() as cursor:
        try:
            cursor.execute("""
                SELECT day_of_week,
                       ISNULL(is_open,1) AS is_open,
                       CONVERT(VARCHAR(5), start_time, 108) AS st,
                       CONVERT(VARCHAR(5), end_time, 108)   AS et
                  FROM OPERATING_TIME
            """)
            for dow, openflag, st, et in cursor.fetchall():
                existing[int(dow)] = {
                    "is_open": bool(openflag),
                    "open": st or default_open,
                    "close": et or default_close,
                }
        except Exception:
            cursor.execute("SELECT day_of_week, is_open, start_time, end_time FROM OPERATING_TIME")
            for dow, openflag, st, et in cursor.fetchall():
                existing[int(dow)] = {
                    "is_open": bool(openflag if openflag is not None else 1),
                    "open": _time_hhmm(st) or default_open,
                    "close": _time_hhmm(et) or default_close,
                }

    week = []
    for dow, label in days_order:
        data = existing.get(dow, {"is_open": True, "open": default_open, "close": default_close})
        week.append({
            "dow": dow,
            "label": label,
            "is_open": data["is_open"],
            "open": data["open"],
            "close": data["close"],
        })

    context.update({
        "week": week,
        "default_open": default_open,
        "default_close": default_close,
    })
    return render(request, "admin/manage_operating_time.html", context)


@require_http_methods(["GET", "POST"])
def branding_settings(request):

    if not is_admin_role(request):
        if request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"ok": False, "error": "forbidden"}, status=403)
        messages.error(request, "Only admins can edit branding.")
        return redirect("/")

    s = SiteSetting.get_solo()

    if request.method == "POST":
        form = SiteSettingForm(request.POST, request.FILES, instance=s)
        if form.is_valid():
            s = form.save()

            try:
                from django.core.cache import cache
                cache.delete("branding_ctx")
            except Exception:
                pass

  
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse({
                    "ok": True,
                    "brand_name": s.brand_name,
                    "brand_logo_url": (s.brand_logo.url if s.brand_logo else None),
                })


            messages.success(request, "Branding updated.")
            next_url = request.POST.get("next") or request.META.get("HTTP_REFERER") or "/"
            return redirect(next_url)

    else:
        form = SiteSettingForm(instance=s)

  
    return render(request, "admin/branding_settings.html", {"form": form})

def _push_active_sessions_for_lab(cursor, lab_id):
    """
    Re-query active sessions for this lab and push via Pusher.
    Frontend can subscribe to channel 'lab-<lab_id>' (optional).
    """
    cursor.execute("""
        SELECT date, start_time, end_time, requested_by
        FROM UTILIZATION_SLIP
        WHERE lab_id = %s
          AND status = 'Active'
          AND date = %s
    """, [lab_id, timezone.localdate()])
    rows = cursor.fetchall()

    payload = {
        "active_sessions": [
            {
                "date":       r[0].isoformat(),
                "start_time": r[1].strftime("%H:%M:%S") if r[1] else None,
                "end_time":   r[2].strftime("%H:%M:%S") if r[2] else None,
                "faculty_id": r[3],
            }
            for r in rows
        ]
    }

    # Optional realtime channel for live highlighting
    try:
        realtime.trigger(f"lab-{lab_id}", "active_sessions", payload)
    except Exception:
        # Avoid breaking RFID flow if Pusher fails
        pass
def _complete_overdue_sessions(cursor, today, now_t):
    """
    Auto-complete any 'Active' sessions that should already be finished
    (end_time < now). Also tap-out any remaining students and push
    updated active_sessions for affected labs.
    """
    from datetime import datetime as dt

    cursor.execute("""
        SELECT utilization_id, lab_id, start_time, end_time
        FROM UTILIZATION_SLIP
        WHERE status = 'Active'
          AND date   = %s
          AND end_time < %s
    """, [today, now_t])
    rows = cursor.fetchall()

    if not rows:
        return

    affected_labs = set()

    for util_id, lab_id, st, et in rows:
        affected_labs.add(lab_id)

        if st and et:
            duration = dt.combine(today, et) - dt.combine(today, st)
            time_dur = (dt.min + duration).time()
        else:
            time_dur = None

        # Mark slip completed
        cursor.execute("""
            UPDATE UTILIZATION_SLIP
               SET status='Completed',
                   time_duration=%s
             WHERE utilization_id=%s
        """, [time_dur, util_id])

        # Auto tap-out students still inside
        cursor.execute("""
            UPDATE COMPUTER_LAB_ATTENDANCE
               SET time_out = %s
             WHERE utilization_id = %s
               AND time_out IS NULL
        """, [et or now_t, util_id])

    # Push updated active_sessions for each affected lab
    for lab_id in affected_labs:
        _push_active_sessions_for_lab(cursor, lab_id)
